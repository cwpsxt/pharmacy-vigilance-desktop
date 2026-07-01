import os
import pandas as pd
import re
import uuid
from datetime import datetime, timedelta
from flask import Blueprint, request, jsonify, session, current_app
from werkzeug.utils import secure_filename
from .db import db
from .models import User, AdverseReactionReport, ImportHistory
from .auth import require_auth

data_bp = Blueprint("data", __name__)

def draw_pie_with_labels(ax, values, labels, colors, title):
    """绘制带引导线标签的饼图，自动避免重叠"""
    import numpy as np
    import textwrap
    
    wedges, texts, autotexts = ax.pie(values, autopct='%1.1f%%', 
        colors=colors[:len(values)], startangle=45, pctdistance=0.75,
        wedgeprops=dict(width=0.5, edgecolor='white', linewidth=2))
    
    bbox_props = dict(boxstyle="round,pad=0.3", fc="white", ec="gray", lw=0.5, alpha=0.9)
    
    # 分左右两边
    left_labels = []
    right_labels = []
    for i, p in enumerate(wedges):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        x = np.cos(np.deg2rad(ang))
        if x < 0:
            left_labels.append((i, ang))
        else:
            right_labels.append((i, ang))
    
    # 给左右两边的标签分配均匀的y坐标
    def draw_side_labels(label_list, side):
        n = len(label_list)
        if n == 0:
            return
        y_positions = np.linspace(0.9, -0.9, n) if n > 1 else [0]
        sorted_list = sorted(label_list, key=lambda x: -np.sin(np.deg2rad(x[1])))
        for idx, (i, ang) in enumerate(sorted_list):
            y_orig = np.sin(np.deg2rad(ang))
            x_orig = np.cos(np.deg2rad(ang))
            x_text = 1.6 if side == 'right' else -1.6
            y_text = y_positions[idx]
            ha = 'left' if side == 'right' else 'right'
            # 长文本换行显示
            label_text = '\n'.join(textwrap.wrap(labels[i], width=25))
            ax.annotate(label_text, xy=(x_orig*0.7, y_orig*0.7), xytext=(x_text, y_text),
                horizontalalignment=ha, va="center", fontsize=9,
                arrowprops=dict(arrowstyle="-", color='gray', connectionstyle="arc3,rad=0"),
                bbox=bbox_props)
    
    draw_side_labels(right_labels, 'right')
    draw_side_labels(left_labels, 'left')
    
    ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
    ax.set_xlim(-2.2, 2.2)
    ax.set_ylim(-1.3, 1.3)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def apply_month_filter(query, start_month='', end_month=''):
    """应用月份区间筛选到查询"""
    if start_month and end_month:
        query = query.filter(
            db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) >= start_month,
            db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) <= end_month
        )
    elif start_month:
        query = query.filter(
            db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) >= start_month
        )
    elif end_month:
        query = query.filter(
            db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) <= end_month
        )
    return query

def get_month_range_str(start_month='', end_month=''):
    """获取月份区间的字符串表示，用于文件名"""
    if start_month and end_month:
        return f"{start_month}_至_{end_month}"
    elif start_month:
        return f"{start_month}_起"
    elif end_month:
        return f"至_{end_month}"
    else:
        return "全部"

IMPORT_REQUIRED_COLUMNS = [
	'报告表编码',
	'报告类型-新的',
	'报告类型-严重程度',
	'病历号/门诊号',
	'怀疑/并用',
	'生产厂家',
	'不良反应名称',
	'报告人职业',
	'报告人签名',
	'国家中心接收时间'
]

IMPORT_OPTIONAL_COLUMNS = ['通用名称']

def normalize_excel_column_name(value):
	if value is None or pd.isna(value):
		return ''
	return str(value).replace('\n', '').replace('\r', '').replace('\u3000', '').strip()

def normalize_import_dataframe(df):
	df = df.copy()
	df.columns = [normalize_excel_column_name(col) for col in df.columns]
	df = df.loc[:, [col for col in df.columns if col]]
	df = df.dropna(how='all')
	for column in IMPORT_OPTIONAL_COLUMNS:
		if column not in df.columns:
			df[column] = ''
	missing = [column for column in IMPORT_REQUIRED_COLUMNS if column not in df.columns]
	if missing:
		raise ValueError(f"缺少必要列: {', '.join(missing)}")
	return df

def read_import_excel(filepath):
	"""读取中心导出或模板Excel，自动识别表头行"""
	raw_sheets = pd.read_excel(filepath, sheet_name=None, header=None)
	for raw_df in raw_sheets.values():
		max_scan_rows = min(len(raw_df), 20)
		for row_index in range(max_scan_rows):
			headers = [normalize_excel_column_name(value) for value in raw_df.iloc[row_index].tolist()]
			if all(column in headers for column in IMPORT_REQUIRED_COLUMNS):
				data = raw_df.iloc[row_index + 1:].copy()
				data.columns = headers
				return normalize_import_dataframe(data)
	return normalize_import_dataframe(pd.read_excel(filepath))

def clean_cell(row, column, default=''):
	value = row.get(column, default)
	if pd.isna(value):
		return default
	return str(value).strip()

def get_deduplicated_records(start_month='', end_month=''):
	query = AdverseReactionReport.query
	query = apply_month_filter(query, start_month, end_month)
	filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
	subquery = filtered_query.with_entities(
		AdverseReactionReport.report_code,
		db.func.min(AdverseReactionReport.id).label('min_id')
	).group_by(AdverseReactionReport.report_code).subquery()
	return db.session.query(AdverseReactionReport).join(
		subquery, AdverseReactionReport.id == subquery.c.min_id
	)

def split_reaction_names(value):
	if not value or pd.isna(value):
		return ['未知不良反应']
	names = []
	for part in re.split(r'[；;、,，]+', str(value)):
		name = re.sub(r'[（(]\s*(一般|严重)\s*[）)]', '', part).strip()
		if name:
			names.append(name)
	return names or ['未知不良反应']

def build_reaction_summary(records):
	reaction_stats = {}
	for record in records:
		for reaction_name in split_reaction_names(record.adverse_reaction_name):
			if reaction_name not in reaction_stats:
				reaction_stats[reaction_name] = {'reaction_name': reaction_name, '严重': 0, '一般': 0, 'total': 0}
			if record.severity == '严重':
				reaction_stats[reaction_name]['严重'] += 1
			else:
				reaction_stats[reaction_name]['一般'] += 1
			reaction_stats[reaction_name]['total'] += 1
	reaction_list = list(reaction_stats.values())
	reaction_list.sort(key=lambda x: x['total'], reverse=True)
	return reaction_list

def build_reaction_summary_excel_data(records):
	reaction_list = build_reaction_summary(records)
	excel_data = []
	for i, item in enumerate(reaction_list, 1):
		excel_data.append({
			'序号': i,
			'不良反应名称': item['reaction_name'],
			'严重': item['严重'],
			'一般': item['一般'],
			'合计': item['total']
		})
	if excel_data:
		total_severe = sum(item['严重'] for item in excel_data)
		total_general = sum(item['一般'] for item in excel_data)
		excel_data.append({
			'序号': '',
			'不良反应名称': '合计',
			'严重': total_severe,
			'一般': total_general,
			'合计': total_severe + total_general
		})
	return excel_data

def add_reaction_summary_charts(writer, records, sheet_name='不良反应汇总分析'):
	reaction_list = build_reaction_summary(records)
	if not reaction_list:
		return
	try:
		from io import BytesIO
		import matplotlib
		matplotlib.use('Agg')
		import matplotlib.pyplot as plt
		import numpy as np
		from openpyxl.drawing.image import Image

		plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
		plt.rcParams['axes.unicode_minus'] = False

		worksheet = writer.sheets[sheet_name]
		top_reactions = reaction_list[:10]
		reaction_names = [item['reaction_name'][:15] + '..' if len(item['reaction_name']) > 15 else item['reaction_name'] for item in top_reactions]
		general_counts = [item['一般'] for item in top_reactions]
		severe_counts = [item['严重'] for item in top_reactions]

		fig1, ax1 = plt.subplots(figsize=(12, 7))
		x = np.arange(len(reaction_names))
		width = 0.35
		ax1.barh(x - width / 2, general_counts[::-1], width, label='一般', color='#67C23A')
		ax1.barh(x + width / 2, severe_counts[::-1], width, label='严重', color='#F56C6C')
		ax1.set_yticks(x)
		ax1.set_yticklabels(reaction_names[::-1], fontsize=9)
		ax1.set_title('不良反应汇总TOP10', fontsize=14, fontweight='bold')
		ax1.set_xlabel('数量')
		ax1.legend()
		plt.tight_layout()
		bar_image = BytesIO()
		plt.savefig(bar_image, format='png', dpi=150, bbox_inches='tight')
		bar_image.seek(0)
		plt.close(fig1)

		img1 = Image(bar_image)
		img1.width = 600
		img1.height = 350
		worksheet.add_image(img1, 'G2')

		labels = [f"{item['reaction_name']}: {item['total']}例" for item in top_reactions]
		values = [item['total'] for item in top_reactions]
		other_total = sum(item['total'] for item in reaction_list[10:])
		if other_total > 0:
			labels.append(f'其他: {other_total}例')
			values.append(other_total)

		fig2, ax2 = plt.subplots(figsize=(12, 9))
		colors = ['#409EFF', '#67C23A', '#E6A23C', '#F56C6C', '#909399', '#606266', '#E91E63', '#9C27B0', '#3F51B5', '#00BCD4', '#CCCCCC']
		draw_pie_with_labels(ax2, values, labels, colors, '不良反应分布')
		pie_image = BytesIO()
		plt.savefig(pie_image, format='png', dpi=150, bbox_inches='tight')
		pie_image.seek(0)
		plt.close(fig2)

		img2 = Image(pie_image)
		img2.width = 600
		img2.height = 450
		worksheet.add_image(img2, 'G24')
	except Exception as chart_error:
		print(f"生成不良反应汇总图表时出错: {chart_error}")

def style_template_header(worksheet):
	from openpyxl.styles import Alignment, Font, PatternFill
	header_fill = PatternFill("solid", fgColor="EAF3FF")
	for row in worksheet.iter_rows(min_row=1, max_row=2):
		for cell in row:
			cell.alignment = Alignment(horizontal="center", vertical="center")
			cell.font = Font(bold=True)
			cell.fill = header_fill
	worksheet.freeze_panes = "A3"

def apply_report_details_template(worksheet):
	for column, title in [('A', '科室'), ('B', '职业'), ('C', '姓名'), ('D', '总计')]:
		worksheet.merge_cells(f'{column}1:{column}2')
		worksheet[f'{column}1'] = title
	worksheet.merge_cells('E1:H1')
	worksheet['E1'] = '报告类型'
	for cell, title in [('E2', '一般'), ('F2', '严重'), ('G2', '新的一般'), ('H2', '新的严重')]:
		worksheet[cell] = title
	style_template_header(worksheet)

def apply_reward_template(worksheet):
	for column, title in [('A', '科室'), ('B', '职业'), ('C', '姓名'), ('D', '总数量'), ('M', '个人奖励合计/元')]:
		worksheet.merge_cells(f'{column}1:{column}2')
		worksheet[f'{column}1'] = title
	for start, end, title in [('E', 'F', '一般'), ('G', 'H', '严重'), ('I', 'J', '新的一般'), ('K', 'L', '新的严重')]:
		worksheet.merge_cells(f'{start}1:{end}1')
		worksheet[f'{start}1'] = title
	for cell, title in [('E2', '数量'), ('F2', '奖励/元'), ('G2', '数量'), ('H2', '奖励/元'), ('I2', '数量'), ('J2', '奖励/元'), ('K2', '数量'), ('L2', '奖励/元')]:
		worksheet[cell] = title
	style_template_header(worksheet)

def parse_excel_data(df, batch_id):
	"""解析Excel数据并转换为数据库记录"""
	records = []
	failed_records = []
	
	# 列名映射
	column_mapping = {
		'报告表编码': 'report_code',
		'报告类型-新的': 'report_type_new',
		'报告类型-严重程度': 'severity',
		'病历号/门诊号': 'medical_record_no',
		'怀疑/并用': 'suspect_concurrent',
		'通用名称': 'generic_name',
		'生产厂家': 'manufacturer',
		'不良反应名称': 'adverse_reaction_name',
		'报告人职业': 'reporter_profession',
		'报告人签名': 'reporter_signature',
		'国家中心接收时间': 'national_center_receive_time'
	}
	
	for index, row in df.iterrows():
		try:
			report_code = clean_cell(row, '报告表编码')
			if not report_code:
				continue
			# 转换时间格式
			receive_time = pd.to_datetime(row['国家中心接收时间'], errors='coerce')
			if pd.isna(receive_time):
				raise ValueError("国家中心接收时间格式不正确")
			
			# 创建记录对象
			record = AdverseReactionReport(
				report_code=report_code,
				report_type_new=clean_cell(row, '报告类型-新的') or None,
				severity=clean_cell(row, '报告类型-严重程度'),
				medical_record_no=clean_cell(row, '病历号/门诊号'),
				suspect_concurrent=clean_cell(row, '怀疑/并用'),
				generic_name=clean_cell(row, '通用名称'),
				manufacturer=clean_cell(row, '生产厂家'),
				adverse_reaction_name=clean_cell(row, '不良反应名称'),
				reporter_profession=clean_cell(row, '报告人职业'),
				reporter_signature=clean_cell(row, '报告人签名'),
				national_center_receive_time=receive_time,
				import_batch_id=batch_id
			)
			records.append(record)
			
		except Exception as e:
			failed_records.append({
				'row': index + 1,
				'error': str(e),
				'data': row.to_dict()
			})
	
	return records, failed_records

@data_bp.route("/upload", methods=["POST"])
@require_auth
def upload_file():
	"""处理Excel文件上传和数据导入"""
	if 'file' not in request.files:
		return jsonify({"message": "没有文件"}), 400
	
	file = request.files['file']
	if file.filename == '':
		return jsonify({"message": "没有选择文件"}), 400
	
	if file and allowed_file(file.filename):
		original_filename = file.filename
		filename = secure_filename(file.filename)
		# 添加时间戳避免文件名冲突
		timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_")
		filename = timestamp + filename
		
		# 确保上传目录存在（使用 Flask config 中配置的可写路径）
		upload_path = current_app.config.get("UPLOAD_FOLDER")
		os.makedirs(upload_path, exist_ok=True)
		
		filepath = os.path.join(upload_path, filename)
		file.save(filepath)
		
		# 生成批次ID
		batch_id = str(uuid.uuid4())
		
		# 创建导入历史记录
		import_history = ImportHistory(
			batch_id=batch_id,
			filename=filename,
			original_filename=original_filename,
			total_records=0,
			success_records=0,
			failed_records=0,
			status='processing'
		)
		db.session.add(import_history)
		db.session.commit()
		
		try:
			# 读取Excel文件
			df = read_import_excel(filepath)
			total_records = len(df)
			
			# 更新总记录数
			import_history.total_records = total_records
			db.session.commit()
			
			# 解析Excel数据
			records, failed_records = parse_excel_data(df, batch_id)
			
			# 批量插入成功的记录
			if records:
				db.session.add_all(records)
			
			# 更新导入历史
			import_history.success_records = len(records)
			import_history.failed_records = len(failed_records)
			import_history.status = 'success' if len(failed_records) == 0 else 'partial_success'
			import_history.completed_at = datetime.utcnow()
			
			if failed_records:
				import_history.error_message = f"部分记录导入失败: {len(failed_records)} 条"
			
			db.session.commit()
			
			return jsonify({
				"message": "文件导入成功",
				"batch_id": batch_id,
				"filename": filename,
				"total_records": total_records,
				"success_records": len(records),
				"failed_records": len(failed_records),
				"upload_time": datetime.now().isoformat()
			})
		
		except Exception as e:
			# 更新导入历史为失败状态
			import_history.status = 'failed'
			import_history.error_message = str(e)
			import_history.completed_at = datetime.utcnow()
			db.session.commit()
			
			# 删除上传失败的文件
			if os.path.exists(filepath):
				os.remove(filepath)
			return jsonify({"message": f"文件处理失败: {str(e)}"}), 400
	
	return jsonify({"message": "不支持的文件格式"}), 400

@data_bp.route("/import-history", methods=["GET"])
@require_auth
def get_import_history():
	"""获取导入历史"""
	try:
		history_records = ImportHistory.query.order_by(ImportHistory.created_at.desc()).limit(50).all()
		return jsonify([record.to_dict() for record in history_records])
	except Exception as e:
		return jsonify({"message": f"获取导入历史失败: {str(e)}"}), 500

@data_bp.route("/import-history/<int:history_id>", methods=["DELETE"])
@require_auth
def delete_import_history(history_id):
	"""删除单次导入及其对应报告"""
	try:
		history = db.session.get(ImportHistory, history_id)
		if not history:
			return jsonify({"message": "导入记录不存在"}), 404

		deleted_count = AdverseReactionReport.query.filter_by(
			import_batch_id=history.batch_id
		).delete(synchronize_session=False)

		upload_path = current_app.config.get("UPLOAD_FOLDER")
		if upload_path and history.filename:
			filepath = os.path.join(upload_path, history.filename)
			if os.path.exists(filepath):
				os.remove(filepath)

		db.session.delete(history)
		db.session.commit()

		return jsonify({
			"message": f"已删除导入记录及 {deleted_count} 条报告",
			"count": deleted_count
		})
	except Exception as e:
		db.session.rollback()
		return jsonify({"message": f"删除导入记录失败: {str(e)}"}), 500

@data_bp.route("/download-template", methods=["GET"])
@require_auth
def download_template():
	"""下载导入模板"""
	try:
		from io import BytesIO
		from flask import send_file
		
		# 创建模板数据
		template_data = {
			'报告表编码': ['示例：ADR202401001'],
			'报告类型-新的': ['新的 或 空'],
			'报告类型-严重程度': ['严重 或 一般'],
			'病历号/门诊号': ['示例：1234567'],
			'怀疑/并用': ['怀疑 或 并用'],
			'通用名称': ['示例：阿莫西林'],
			'生产厂家': ['示例：XX制药有限公司'],
			'不良反应名称': ['示例：皮疹'],
			'报告人职业': ['示例：药师'],
			'报告人签名': ['示例：张三'],
			'国家中心接收时间': ['示例：2024-01-15 10:30:00']
		}
		
		df = pd.DataFrame(template_data)
		
		output = BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name='数据导入模板', index=False)
		
		output.seek(0)
		
		return send_file(
			output,
			as_attachment=True,
			download_name='药品不良反应数据导入模板.xlsx',
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		
	except Exception as e:
		return jsonify({"message": f"下载模板失败: {str(e)}"}), 500

@data_bp.route("/stats", methods=["GET"])
@require_auth
def get_stats():
	"""获取仪表板统计数据"""
	try:
		# 总记录数
		total_records = AdverseReactionReport.query.count()
		
		# 今日导入数
		today = datetime.now().date()
		today_imports = ImportHistory.query.filter(
			db.func.date(ImportHistory.created_at) == today
		).with_entities(db.func.sum(ImportHistory.success_records)).scalar() or 0
		
		# 去重数据总数（只统计"怀疑"的数据并按报告编码去重）
		filtered_query = AdverseReactionReport.query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_count = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		).count()
		
		# 严重不良反应数量（作为风险警报）
		serious_alerts = AdverseReactionReport.query.filter_by(severity='严重').count()
		
		# 本月新增数据量
		current_month = datetime.now().strftime('%Y-%m')
		monthly_count = AdverseReactionReport.query.filter(
			db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) == current_month
		).count()
		
		# 药师上报数量
		pharmacist_count = AdverseReactionReport.query.filter(
			AdverseReactionReport.reporter_profession == '药师'
		).count()
		
		# 新的报告数量
		new_reports = AdverseReactionReport.query.filter(
			AdverseReactionReport.report_type_new == '新的'
		).count()
		
		# 最近7天导入趋势
		import_trend = []
		for i in range(7):
			date = (datetime.now() - timedelta(days=i)).date()
			daily_imports = ImportHistory.query.filter(
				db.func.date(ImportHistory.created_at) == date
			).with_entities(db.func.sum(ImportHistory.success_records)).scalar() or 0
			import_trend.append({
				'date': date.strftime('%m-%d'),
				'count': int(daily_imports)
			})
		import_trend.reverse()  # 按时间正序
		
		# 按严重程度分布
		severity_distribution = []
		for severity in ['一般', '严重']:
			count = AdverseReactionReport.query.filter_by(severity=severity).count()
			percentage = (count / total_records * 100) if total_records > 0 else 0
			severity_distribution.append({
				'name': severity,
				'count': count,
				'percentage': round(percentage, 1)
			})
		
		# 按科室分布（Top 5）
		department_stats = db.session.query(
			db.case(
				(AdverseReactionReport.reporter_profession.like('%GCP%'), 'GCP'),
				else_='药剂科'
			).label('department'),
			db.func.count(AdverseReactionReport.id).label('count')
		).group_by('department').order_by(db.func.count(AdverseReactionReport.id).desc()).limit(5).all()
		
		department_distribution = [
			{'name': dept[0], 'count': dept[1]} 
			for dept in department_stats
		]
		
		stats = {
			"totalRecords": total_records,
			"todayImports": int(today_imports),
			"pendingReviews": deduplicated_count,  # 改为显示去重数据数量
			"alerts": serious_alerts,
			"monthlyCount": monthly_count,
			"pharmacistCount": pharmacist_count,
			"newReports": new_reports,
			"importTrend": import_trend,
			"severityDistribution": severity_distribution,
			"departmentDistribution": department_distribution
		}
		return jsonify(stats)
	except Exception as e:
		return jsonify({"message": f"获取统计数据失败: {str(e)}"}), 500

@data_bp.route("/reports", methods=["GET"])
@require_auth
def get_reports():
	"""获取不良反应报告列表"""
	try:
		page = request.args.get('page', 1, type=int)
		per_page = request.args.get('per_page', 20, type=int)
		
		# 查询参数
		search = request.args.get('search', '', type=str)
		severity = request.args.get('severity', '', type=str)
		suspect_concurrent = request.args.get('suspect_concurrent', '', type=str)
		start_month = request.args.get('startMonth', '', type=str)
		end_month = request.args.get('endMonth', '', type=str)
		
		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 搜索过滤
		if search:
			query = query.filter(
				db.or_(
					AdverseReactionReport.report_code.contains(search),
					AdverseReactionReport.generic_name.contains(search),
					AdverseReactionReport.manufacturer.contains(search),
					AdverseReactionReport.adverse_reaction_name.contains(search)
				)
			)
		
		# 严重程度过滤
		if severity:
			query = query.filter_by(severity=severity)
		if suspect_concurrent:
			query = query.filter_by(suspect_concurrent=suspect_concurrent)
		
		# 分页
		pagination = query.order_by(AdverseReactionReport.created_at.desc()).paginate(
			page=page, per_page=per_page, error_out=False
		)
		
		return jsonify({
			"reports": [report.to_dict() for report in pagination.items],
			"total": pagination.total,
			"pages": pagination.pages,
			"current_page": page,
			"per_page": per_page
		})
		
	except Exception as e:
		return jsonify({"message": f"获取报告列表失败: {str(e)}"}), 500

@data_bp.route("/analysis/months", methods=["GET"])
@require_auth
def get_available_months():
	"""获取可用的月份列表"""
	try:
		# 查询数据库中所有不同的年月
		months = db.session.query(
			db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time).label('month')
		).distinct().order_by('month').all()
		
		month_list = [{'value': month[0], 'label': month[0]} for month in months if month[0]]
		return jsonify(month_list)
	except Exception as e:
		return jsonify({"message": f"获取月份列表失败: {str(e)}"}), 500

@data_bp.route("/analysis/field-values", methods=["GET"])
@require_auth
def get_field_values():
	"""获取各字段的唯一值列表"""
	try:
		field = request.args.get('field', '')
		if not field:
			return jsonify({"message": "字段参数不能为空"}), 400
		
		# 字段映射
		field_mapping = {
			'severity': AdverseReactionReport.severity,
			'suspect_concurrent': AdverseReactionReport.suspect_concurrent,
			'reporter_profession': AdverseReactionReport.reporter_profession,
			'manufacturer': AdverseReactionReport.manufacturer,
			'generic_name': AdverseReactionReport.generic_name
		}
		
		if field not in field_mapping:
			return jsonify({"message": "不支持的字段"}), 400
		
		# 获取字段的唯一值
		values = db.session.query(field_mapping[field]).distinct().all()
		value_list = [{'value': val[0], 'label': val[0]} for val in values if val[0]]
		
		return jsonify(sorted(value_list, key=lambda x: x['label']))
	except Exception as e:
		return jsonify({"message": f"获取字段值失败: {str(e)}"}), 500

@data_bp.route("/analysis/data", methods=["GET"])
@require_auth
def get_analysis_data():
	"""获取分析数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')
		severity = request.args.get('severity', '')
		suspect_concurrent = request.args.get('suspect_concurrent', '')
		reporter_profession = request.args.get('reporter_profession', '')
		manufacturer = request.args.get('manufacturer', '')
		generic_name = request.args.get('generic_name', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)

		# 按其他字段筛选
		if severity:
			query = query.filter(AdverseReactionReport.severity == severity)
		if suspect_concurrent:
			query = query.filter(AdverseReactionReport.suspect_concurrent == suspect_concurrent)
		if reporter_profession:
			query = query.filter(AdverseReactionReport.reporter_profession == reporter_profession)
		if manufacturer:
			query = query.filter(AdverseReactionReport.manufacturer == manufacturer)
		if generic_name:
			query = query.filter(AdverseReactionReport.generic_name == generic_name)
		
		# 获取筛选后的数据
		reports = query.order_by(AdverseReactionReport.created_at.desc()).all()
		
		return jsonify({
			"reports": [report.to_dict() for report in reports],
			"total": len(reports)
		})
		
	except Exception as e:
		return jsonify({"message": f"获取分析数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/deduplicated", methods=["GET"])
@require_auth
def get_deduplicated_data():
	"""获取去重数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')
		severity = request.args.get('severity', '')
		suspect_concurrent = request.args.get('suspect_concurrent', '')
		reporter_profession = request.args.get('reporter_profession', '')
		manufacturer = request.args.get('manufacturer', '')
		generic_name = request.args.get('generic_name', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)

		# 按其他字段筛选
		if severity:
			query = query.filter(AdverseReactionReport.severity == severity)
		if suspect_concurrent:
			query = query.filter(AdverseReactionReport.suspect_concurrent == suspect_concurrent)
		if reporter_profession:
			query = query.filter(AdverseReactionReport.reporter_profession == reporter_profession)
		if manufacturer:
			query = query.filter(AdverseReactionReport.manufacturer == manufacturer)
		if generic_name:
			query = query.filter(AdverseReactionReport.generic_name == generic_name)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		).order_by(AdverseReactionReport.created_at.desc())
		
		reports = deduplicated_query.all()
		
		return jsonify({
			"reports": [report.to_dict() for report in reports],
			"total": len(reports)
		})
		
	except Exception as e:
		return jsonify({"message": f"获取去重数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/export", methods=["GET"])
@require_auth
def export_analysis_data():
	"""导出分析数据为Excel"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')
		severity = request.args.get('severity', '')
		suspect_concurrent = request.args.get('suspect_concurrent', '')
		reporter_profession = request.args.get('reporter_profession', '')
		manufacturer = request.args.get('manufacturer', '')
		generic_name = request.args.get('generic_name', '')
		export_type = request.args.get('type', 'all')  # all 或 deduplicated

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 按其他字段筛选
		if severity:
			query = query.filter(AdverseReactionReport.severity == severity)
		if suspect_concurrent:
			query = query.filter(AdverseReactionReport.suspect_concurrent == suspect_concurrent)
		if reporter_profession:
			query = query.filter(AdverseReactionReport.reporter_profession == reporter_profession)
		if manufacturer:
			query = query.filter(AdverseReactionReport.manufacturer == manufacturer)
		if generic_name:
			query = query.filter(AdverseReactionReport.generic_name == generic_name)
		
		# 根据导出类型处理数据
		if export_type == 'deduplicated':
			# 先过滤掉"并用"的数据，只保留"怀疑"的数据
			filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
			
			# 然后按报告编码去重
			subquery = filtered_query.with_entities(
				AdverseReactionReport.report_code,
				db.func.min(AdverseReactionReport.id).label('min_id')
			).group_by(AdverseReactionReport.report_code).subquery()
			
			reports_query = db.session.query(AdverseReactionReport).join(
				subquery, AdverseReactionReport.id == subquery.c.min_id
			).order_by(AdverseReactionReport.created_at.desc())
		else:
			reports_query = query.order_by(AdverseReactionReport.created_at.desc())
		
		reports = reports_query.all()
		
		# 转换为DataFrame
		import pandas as pd
		data = []
		for report in reports:
			data.append({
				'报告表编码': report.report_code,
				'报告类型-新的': report.report_type_new or '',
				'报告类型-严重程度': report.severity,
				'病历号/门诊号': report.medical_record_no,
				'怀疑/并用': report.suspect_concurrent,
				'通用名称': report.generic_name,
				'生产厂家': report.manufacturer,
				'不良反应名称': report.adverse_reaction_name,
				'报告人职业': report.reporter_profession,
				'报告人签名': report.reporter_signature,
				'国家中心接收时间': report.national_center_receive_time.strftime('%Y-%m-%d %H:%M:%S') if report.national_center_receive_time else ''
			})
		
		df = pd.DataFrame(data)
		
		# 生成Excel文件
		from io import BytesIO
		from flask import send_file
		
		output = BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name='数据', index=False)
		
		output.seek(0)
		
		# 生成文件名
		type_name = "去重数据" if export_type == 'deduplicated' else "全部数据"
		month_range = get_month_range_str(start_month, end_month)
		filename = f"药品不良反应{type_name}_{month_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			as_attachment=True,
			download_name=filename,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		
	except Exception as e:
		return jsonify({"message": f"导出数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/statistics", methods=["GET"])
@require_auth
def get_statistics():
	"""获取类型统计数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query

		# 按月份区间筛选
		if start_month and end_month:
			query = query.filter(
				db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) >= start_month,
				db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) <= end_month
			)
		elif start_month:
			query = query.filter(
				db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) >= start_month
			)
		elif end_month:
			query = query.filter(
				db.func.strftime('%Y-%m', AdverseReactionReport.national_center_receive_time) <= end_month
			)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有记录用于组合统计
		all_records = deduplicated_query.all()
		
		# 正确的统计逻辑：避免重复计算
		# 初始化计数器
		new_count = 0  # "报告类型-新的"为"新的"的数量
		severity_counts = {'严重': 0, '一般': 0}  # 按严重程度统计（所有记录）
		non_new_severity_counts = {'严重': 0, '一般': 0}  # 非新的按严重程度统计
		new_and_severity_counts = {'严重': 0, '一般': 0}  # 既是新的又是对应严重程度的记录数量
		
		for record in all_records:
			# 判断是否为"新的"
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			# 统计"报告类型-新的"为"新的"的记录
			if is_new:
				new_count += 1
				# 统计既是新的又是对应严重程度的记录
				if record.severity in new_and_severity_counts:
					new_and_severity_counts[record.severity] += 1
			else:
				# 统计非新的按严重程度
				if record.severity in non_new_severity_counts:
					non_new_severity_counts[record.severity] += 1
			
			# 统计所有记录的严重程度
			if record.severity in severity_counts:
				severity_counts[record.severity] += 1
		
		# 组合统计结果（避免重复计算）
		type_counts = {
			'新的+严重': new_count + severity_counts['严重'] - new_and_severity_counts['严重'],  # 新的数量 + 严重数量 - 重复部分
			'新的+一般': new_count + severity_counts['一般'] - new_and_severity_counts['一般'],  # 新的数量 + 一般数量 - 重复部分
			'严重': non_new_severity_counts['严重'],          # 只统计非新的严重
			'一般': non_new_severity_counts['一般']           # 只统计非新的一般
		}
		
		# 转换为前端需要的格式，按指定顺序排列
		severity_data = []
		total_count = len(all_records)  # 总计是去重之后的总数
		
		# 按指定顺序展示：一般 -> 严重 -> 新的+一般 -> 新的+严重
		display_order = ['一般', '严重', '新的+一般', '新的+严重']
		
		for type_name in display_order:
			if type_name in type_counts:
				count = type_counts[type_name]
				percentage = (count / total_count * 100) if total_count > 0 else 0
				severity_data.append({
					'type': type_name,
					'count': count,
					'percentage': round(percentage, 1)
				})
		
		# 月度统计
		month_stats = []
		if start_month or end_month:
			# 如果指定了月份区间，使用已获取的记录（已经过滤）进行统计
			month_records = all_records
			
			# 统计该月的数据
			month_new_count = 0
			month_severity_counts = {'严重': 0, '一般': 0}  # 所有记录的严重程度统计
			month_non_new_severity_counts = {'严重': 0, '一般': 0}  # 非新的记录的严重程度统计
			month_new_and_severity_counts = {'严重': 0, '一般': 0}  # 既是新的又是对应严重程度的记录数量
			
			for record in month_records:
				# 判断是否为"新的"
				is_new = (record.report_type_new and 
						 record.report_type_new.strip() and 
						 record.report_type_new == '新的')
				
				# 统计"报告类型-新的"为"新的"的记录
				if is_new:
					month_new_count += 1
					# 统计既是新的又是对应严重程度的记录
					if record.severity in month_new_and_severity_counts:
						month_new_and_severity_counts[record.severity] += 1
				else:
					# 统计非新的按严重程度
					if record.severity in month_non_new_severity_counts:
						month_non_new_severity_counts[record.severity] += 1
				
				# 统计所有记录的严重程度
				if record.severity in month_severity_counts:
					month_severity_counts[record.severity] += 1
			
			# 计算组合统计结果（避免重复计算）
			month_type_counts = {
				'新的+严重': month_new_count + month_severity_counts['严重'] - month_new_and_severity_counts['严重'],  # 新的数量 + 严重数量 - 重复部分
				'新的+一般': month_new_count + month_severity_counts['一般'] - month_new_and_severity_counts['一般'],  # 新的数量 + 一般数量 - 重复部分
				'严重': month_non_new_severity_counts['严重'],              # 只统计非新的严重
				'一般': month_non_new_severity_counts['一般']               # 只统计非新的一般
			}
			
			# 按月分组统计区间内的数据
			month_data = {}
			for record in month_records:
				month_key = record.national_center_receive_time.strftime('%Y-%m')

				if month_key not in month_data:
					month_data[month_key] = {
						'new_count': 0,
						'severity_counts': {'严重': 0, '一般': 0},
						'non_new_severity_counts': {'严重': 0, '一般': 0},
						'new_and_severity_counts': {'严重': 0, '一般': 0}
					}

				# 判断是否为"新的"
				is_new = (record.report_type_new and
						 record.report_type_new.strip() and
						 record.report_type_new == '新的')

				if is_new:
					month_data[month_key]['new_count'] += 1
					if record.severity in month_data[month_key]['new_and_severity_counts']:
						month_data[month_key]['new_and_severity_counts'][record.severity] += 1
				else:
					if record.severity in month_data[month_key]['non_new_severity_counts']:
						month_data[month_key]['non_new_severity_counts'][record.severity] += 1

				if record.severity in month_data[month_key]['severity_counts']:
					month_data[month_key]['severity_counts'][record.severity] += 1

			# 生成月度统计
			for month_key, data in sorted(month_data.items()):
				new_count = data['new_count']
				severity_counts = data['severity_counts']
				new_and_severity_counts = data['new_and_severity_counts']
				non_new_severity_counts = data['non_new_severity_counts']

				month_type_counts = {
					'新的+严重': new_count + severity_counts['严重'] - new_and_severity_counts['严重'],
					'新的+一般': new_count + severity_counts['一般'] - new_and_severity_counts['一般'],
					'严重': non_new_severity_counts['严重'],
					'一般': non_new_severity_counts['一般']
				}

				month_stats.append({
					'month': month_key,
					'一般': month_type_counts['一般'],
					'严重': month_type_counts['严重'],
					'新的+一般': month_type_counts['新的+一般'],
					'新的+严重': month_type_counts['新的+严重']
				})
		else:
			# 按月份分组统计所有记录
			month_data = {}
			
			for record in all_records:
				month_key = record.national_center_receive_time.strftime('%Y-%m')
				
				if month_key not in month_data:
					month_data[month_key] = {
						'new_count': 0,
						'severity_counts': {'严重': 0, '一般': 0},  # 所有记录的严重程度统计
						'non_new_severity_counts': {'严重': 0, '一般': 0},  # 非新的记录的严重程度统计
						'new_and_severity_counts': {'严重': 0, '一般': 0}  # 既是新的又是对应严重程度的记录数量
					}
				
				# 判断是否为"新的"
				is_new = (record.report_type_new and 
						 record.report_type_new.strip() and 
						 record.report_type_new == '新的')
				
				# 统计"报告类型-新的"为"新的"的记录
				if is_new:
					month_data[month_key]['new_count'] += 1
					# 统计既是新的又是对应严重程度的记录
					if record.severity in month_data[month_key]['new_and_severity_counts']:
						month_data[month_key]['new_and_severity_counts'][record.severity] += 1
				else:
					# 统计非新的按严重程度
					if record.severity in month_data[month_key]['non_new_severity_counts']:
						month_data[month_key]['non_new_severity_counts'][record.severity] += 1
				
				# 统计所有记录的严重程度
				if record.severity in month_data[month_key]['severity_counts']:
					month_data[month_key]['severity_counts'][record.severity] += 1
			
			# 转换为列表格式
			for month_key, data in sorted(month_data.items()):
				new_count = data['new_count']
				severity_counts = data['severity_counts']
				non_new_severity_counts = data['non_new_severity_counts']
				new_and_severity_counts = data['new_and_severity_counts']
				
				month_type_counts = {
					'新的+严重': new_count + severity_counts['严重'] - new_and_severity_counts['严重'],  # 新的数量 + 严重数量 - 重复部分
					'新的+一般': new_count + severity_counts['一般'] - new_and_severity_counts['一般'],  # 新的数量 + 一般数量 - 重复部分
					'严重': non_new_severity_counts['严重'],          # 只统计非新的严重
					'一般': non_new_severity_counts['一般']           # 只统计非新的一般
				}
				
				month_stats.append({
					'month': month_key,
					'一般': month_type_counts['一般'],
					'严重': month_type_counts['严重'],
					'新的+一般': month_type_counts['新的+一般'],
					'新的+严重': month_type_counts['新的+严重']
				})
		
		# 类型分布图数据，只显示总计、严重、新的+严重
		chart_data = []
		# 添加总计
		chart_data.append({
			'name': '总计',
			'value': total_count
		})
		# 只添加严重和新的+严重
		for item in severity_data:
			if item['type'] in ['严重', '新的+严重']:
				chart_data.append({
					'name': item['type'],
					'value': item['count']
				})
		
		# 新增：按月占比统计（使用去重数据）
		monthly_ratio_stats = []
		
		# 获取去重数据进行按月统计
		deduplicated_all_query = AdverseReactionReport.query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		deduplicated_all_query = apply_month_filter(deduplicated_all_query, start_month, end_month)
		
		# 按报告编码去重（取每个报告编码的第一条记录）
		subquery_all = deduplicated_all_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_all_data = db.session.query(AdverseReactionReport).join(
			subquery_all, AdverseReactionReport.id == subquery_all.c.min_id
		).all()
		
		# 按月分组统计去重数据
		monthly_all_data = {}
		for record in deduplicated_all_data:
			month_key = record.national_center_receive_time.strftime('%Y%m')
			
			if month_key not in monthly_all_data:
				monthly_all_data[month_key] = {
					'total': 0,
					'严重的': 0,
					'新的': 0
				}
			
			monthly_all_data[month_key]['total'] += 1
			
			# 统计严重的
			if record.severity == '严重':
				monthly_all_data[month_key]['严重的'] += 1
			
			# 统计新的
			if (record.report_type_new and 
				record.report_type_new.strip() and 
				record.report_type_new == '新的'):
				monthly_all_data[month_key]['新的'] += 1
		
		# 转换为前端需要的格式
		for month_key in sorted(monthly_all_data.keys()):
			data = monthly_all_data[month_key]
			total = data['total']
			
			# 严重的+新的 = 严重的数量 + 新的数量
			serious_plus_new_count = data['严重的'] + data['新的']
			
			serious_ratio = (data['严重的'] / total * 100) if total > 0 else 0
			new_serious_ratio = (serious_plus_new_count / total * 100) if total > 0 else 0
			
			monthly_ratio_stats.append({
				'month': month_key,
				'total': total,
				'serious_count': data['严重的'],
				'serious_ratio': round(serious_ratio, 2),
				'new_serious_count': serious_plus_new_count,
				'new_serious_ratio': round(new_serious_ratio, 2)
			})
		
		return jsonify({
			'severity_stats': severity_data,
			'month_stats': month_stats,
			'chart_data': chart_data,
			'total_count': total_count,
			'monthly_ratio_stats': monthly_ratio_stats
		})
		
	except Exception as e:
		return jsonify({"message": f"获取统计数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/export-statistics", methods=["GET"])
@require_auth
def export_statistics():
	"""导出类型统计数据为Excel"""
	try:
		import pandas as pd
		from io import BytesIO
		from flask import send_file
		
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有记录用于组合统计
		all_records = deduplicated_query.all()
		
		# 正确的统计逻辑：避免重复计算
		# 初始化计数器
		new_count = 0  # "报告类型-新的"为"新的"的数量
		severity_counts = {'严重': 0, '一般': 0}  # 按严重程度统计（所有记录）
		non_new_severity_counts = {'严重': 0, '一般': 0}  # 非新的按严重程度统计
		new_and_severity_counts = {'严重': 0, '一般': 0}  # 既是新的又是对应严重程度的记录数量
		
		for record in all_records:
			# 判断是否为"新的"
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			# 统计"报告类型-新的"为"新的"的记录
			if is_new:
				new_count += 1
				# 统计既是新的又是对应严重程度的记录
				if record.severity in new_and_severity_counts:
					new_and_severity_counts[record.severity] += 1
			else:
				# 统计非新的按严重程度
				if record.severity in non_new_severity_counts:
					non_new_severity_counts[record.severity] += 1
			
			# 统计所有记录的严重程度
			if record.severity in severity_counts:
				severity_counts[record.severity] += 1
		
		# 组合统计结果（避免重复计算）
		type_counts = {
			'新的+严重': new_count + severity_counts['严重'] - new_and_severity_counts['严重'],  # 新的数量 + 严重数量 - 重复部分
			'新的+一般': new_count + severity_counts['一般'] - new_and_severity_counts['一般'],  # 新的数量 + 一般数量 - 重复部分
			'严重': non_new_severity_counts['严重'],          # 只统计非新的严重
			'一般': non_new_severity_counts['一般']           # 只统计非新的一般
		}
		
		# 按指定顺序展示：一般 -> 严重 -> 新的+一般 -> 新的+严重
		display_order = ['一般', '严重', '新的+一般', '新的+严重']
		
		# 创建类型统计数据
		severity_data = []
		total_count = len(all_records)  # 总计是去重之后的总数
		
		for type_name in display_order:
			if type_name in type_counts:
				count = type_counts[type_name]
				percentage = (count / total_count * 100) if total_count > 0 else 0
				severity_data.append({
					'报告类型': type_name,
					'数量': count,
					'占比(%)': round(percentage, 1)
				})
		
		# 获取去重数据进行按月统计
		deduplicated_all_query = AdverseReactionReport.query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		deduplicated_all_query = apply_month_filter(deduplicated_all_query, start_month, end_month)
		
		# 按报告编码去重（取每个报告编码的第一条记录）
		subquery_all = deduplicated_all_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_all_data = db.session.query(AdverseReactionReport).join(
			subquery_all, AdverseReactionReport.id == subquery_all.c.min_id
		).all()
		
		# 按月分组统计去重数据
		monthly_all_data = {}
		for record in deduplicated_all_data:
			month_key = record.national_center_receive_time.strftime('%Y%m')
			
			if month_key not in monthly_all_data:
				monthly_all_data[month_key] = {
					'total': 0,
					'严重的': 0,
					'新的': 0
				}
			
			monthly_all_data[month_key]['total'] += 1
			
			# 统计严重的
			if record.severity == '严重':
				monthly_all_data[month_key]['严重的'] += 1
			
			# 统计新的
			if (record.report_type_new and 
				record.report_type_new.strip() and 
				record.report_type_new == '新的'):
				monthly_all_data[month_key]['新的'] += 1
		
		# 创建月度占比统计数据
		monthly_ratio_data = []
		
		# 生成12个月份的数据
		months = []
		for i in range(1, 13):
			month_key = f"2025{i:02d}"
			months.append(month_key)
		
		# 创建表格数据
		total_row = ['总计'] + [monthly_all_data.get(m, {}).get('total', 0) for m in months]
		serious_row = ['严重的'] + [monthly_all_data.get(m, {}).get('严重的', 0) for m in months]
		serious_ratio_row = ['占比'] + []
		serious_plus_new_row = ['严重的+新的'] + []
		serious_plus_new_ratio_row = ['占比'] + []
		
		# 计算占比和严重的+新的
		total_sum = 0
		serious_sum = 0
		serious_plus_new_sum = 0
		
		for month_key in months:
			data = monthly_all_data.get(month_key, {'total': 0, '严重的': 0, '新的': 0})
			total = data['total']
			serious_count = data['严重的']
			new_count = data['新的']
			serious_plus_new_count = serious_count + new_count
			
			# 计算占比
			serious_ratio = (serious_count / total * 100) if total > 0 else 0
			serious_plus_new_ratio = (serious_plus_new_count / total * 100) if total > 0 else 0
			
			serious_ratio_row.append(f"{serious_ratio:.2f}%")
			serious_plus_new_row.append(serious_plus_new_count)
			serious_plus_new_ratio_row.append(f"{serious_plus_new_ratio:.2f}%")
			
			# 累计总数
			total_sum += total
			serious_sum += serious_count
			serious_plus_new_sum += serious_plus_new_count
		
		# 创建Excel文件
		output = BytesIO()
		
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# 创建一个综合工作表，包含所有数据
			
			# 1. 类型统计表
			df_severity = pd.DataFrame(severity_data)
			df_severity.to_excel(writer, sheet_name='类型统计数据', index=False, startrow=0)
			
			# 2. 空行分隔
			start_row_monthly = len(df_severity) + 3
			
			# 3. ADR数量占比统计表
			columns = ['ADR数量'] + months
			monthly_data = [
				total_row,
				serious_row,
				serious_ratio_row,
				serious_plus_new_row,
				serious_plus_new_ratio_row
			]
			df_monthly = pd.DataFrame(monthly_data, columns=columns)
			df_monthly.to_excel(writer, sheet_name='类型统计数据', index=False, startrow=start_row_monthly)
			
			# 4. 创建柱状图并插入Excel
			try:
				import matplotlib
				matplotlib.use('Agg')  # 使用非交互式后端
				import matplotlib.pyplot as plt
				from matplotlib import font_manager
				import numpy as np
				
				# 设置中文字体，如果没有中文字体就使用默认字体
				try:
					plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
					plt.rcParams['axes.unicode_minus'] = False
				except:
					# 如果字体设置失败，使用默认字体
					pass
				
				# 准备图表数据，只显示总计、严重、新的+严重
				chart_data_filtered = []
				
				# 添加总计
				chart_data_filtered.append({
					'name': '总计',
					'value': total_count
				})
				
				# 只添加严重和新的+严重
				for item in severity_data:
					if item['报告类型'] in ['严重', '新的+严重']:
						chart_data_filtered.append({
							'name': item['报告类型'],
							'value': item['数量']
						})
				
				chart_types = [item['name'] for item in chart_data_filtered]
				chart_values = [item['value'] for item in chart_data_filtered]
				chart_colors = ['#606266', '#F56C6C', '#E6A23C']  # 总计、严重、新的+严重
				
				# 创建图表
				fig, ax = plt.subplots(figsize=(10, 6))
				bars = ax.bar(chart_types, chart_values, color=chart_colors[:len(chart_types)], width=0.5)
				
				# 设置图表样式
				ax.set_title('类型分布图', fontsize=16, fontweight='bold', pad=20)
				ax.set_ylabel('数量', fontsize=12)
				ax.set_xlabel('报告类型', fontsize=12)
				
				# 在柱子上显示数值
				for bar, value in zip(bars, chart_values):
					height = bar.get_height()
					ax.text(bar.get_x() + bar.get_width()/2., height + 0.5,
						   f'{value}', ha='center', va='bottom', fontsize=11, fontweight='bold')
				
				# 设置y轴范围，确保有足够空间显示数值
				max_value = max(chart_values)
				ax.set_ylim(0, max_value * 1.15)
				
				# 美化图表
				ax.grid(True, alpha=0.3, axis='y')
				ax.set_axisbelow(True)
				
				# 调整布局
				plt.tight_layout()
				
				# 保存图表为临时文件
				chart_image = BytesIO()
				plt.savefig(chart_image, format='png', dpi=300, bbox_inches='tight')
				chart_image.seek(0)
				plt.close()
				
				# 获取工作表并插入图片
				worksheet = writer.sheets['类型统计数据']
				
				# 插入图片到右侧空白区域
				from openpyxl.drawing.image import Image
				img = Image(chart_image)
				
				# 调整图片大小
				img.width = 600
				img.height = 360
				
				# 插入图片到合适位置（右侧）
				worksheet.add_image(img, 'H2')
				
			except Exception as chart_error:
				print(f"生成图表时出错: {chart_error}")
				# 如果图表生成失败，继续执行其他部分
		
		output.seek(0)
		
		# 生成文件名
		filename = f"类型统计数据_{get_month_range_str(start_month, end_month)}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=filename
		)
		
	except Exception as e:
		return jsonify({"message": f"导出统计数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/report-details", methods=["GET"])
@require_auth
def get_report_details():
	"""获取上报明细统计数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		# 获取所有去重后的记录
		all_records = get_deduplicated_records(start_month, end_month).all()
		
		# 按报告人统计
		reporter_stats = {}
		
		for record in all_records:
			# 获取科室信息（这里假设从报告人职业或其他字段获取，如果没有科室字段，可以设为默认值）
			department = "药剂科"  # 根据图片，大部分是药剂科，如果有其他科室字段可以替换
			if record.reporter_profession and "GCP" in str(record.reporter_profession):
				department = "GCP"
			
			profession = record.reporter_profession or "药师"
			reporter_name = record.reporter_signature or "未知"
			
			# 创建唯一键
			key = f"{department}|{profession}|{reporter_name}"
			
			if key not in reporter_stats:
				reporter_stats[key] = {
					'department': department,
					'profession': profession,
					'reporter_name': reporter_name,
					'total': 0,
					'一般': 0,
					'严重': 0,
					'新的': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			# 统计总数
			reporter_stats[key]['total'] += 1
			
			# 判断是否是新的报告
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			if is_new:
				reporter_stats[key]['新的'] += 1
				# 新的+严重程度组合
				if record.severity == '一般':
					reporter_stats[key]['新的_一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['新的_严重'] += 1
			else:
				# 非新的记录才统计到一般/严重
				if record.severity == '一般':
					reporter_stats[key]['一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['严重'] += 1
		
		# 转换为列表并计算组合统计
		report_details = []
		total_stats = {
			'total': 0,
			'一般': 0,
			'严重': 0,
			'新的_一般': 0,
			'新的_严重': 0
		}
		
		for stats in reporter_stats.values():
			# 直接使用已经正确计算的新的+严重程度组合统计
			report_detail = {
				'department': stats['department'],
				'profession': stats['profession'],
				'reporter_name': stats['reporter_name'],
				'total': stats['total'],
				'一般': stats['一般'],
				'严重': stats['严重'],
				'新的_一般': stats['新的_一般'],
				'新的_严重': stats['新的_严重']
			}
			
			report_details.append(report_detail)
			
			# 累计总数
			total_stats['total'] += stats['total']
			total_stats['一般'] += stats['一般']
			total_stats['严重'] += stats['严重']
			total_stats['新的_一般'] += stats['新的_一般']
			total_stats['新的_严重'] += stats['新的_严重']
		
		# 按科室、职业、姓名排序
		report_details.sort(key=lambda x: (x['department'], x['profession'], x['reporter_name']))
		
		# 添加小计和合计行
		# 按科室分组计算小计
		department_subtotals = {}
		for detail in report_details:
			dept = detail['department']
			if dept not in department_subtotals:
				department_subtotals[dept] = {
					'total': 0,
					'一般': 0,
					'严重': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			department_subtotals[dept]['total'] += detail['total']
			department_subtotals[dept]['一般'] += detail['一般']
			department_subtotals[dept]['严重'] += detail['严重']
			department_subtotals[dept]['新的_一般'] += detail['新的_一般']
			department_subtotals[dept]['新的_严重'] += detail['新的_严重']
		
		return jsonify({
			'report_details': report_details,
			'department_subtotals': department_subtotals,
			'total_stats': total_stats
		})
		
	except Exception as e:
		return jsonify({"message": f"获取上报明细失败: {str(e)}"}), 500

@data_bp.route("/analysis/export-report-details", methods=["GET"])
@require_auth
def export_report_details():
	"""导出上报明细数据为Excel"""
	try:
		import pandas as pd
		from io import BytesIO
		from flask import send_file
		
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有去重后的记录
		all_records = deduplicated_query.all()
		
		# 按报告人统计
		reporter_stats = {}
		
		for record in all_records:
			# 获取科室信息
			department = "药剂科"
			if record.reporter_profession and "GCP" in str(record.reporter_profession):
				department = "GCP"
			
			profession = record.reporter_profession or "药师"
			reporter_name = record.reporter_signature or "未知"
			
			# 创建唯一键
			key = f"{department}|{profession}|{reporter_name}"
			
			if key not in reporter_stats:
				reporter_stats[key] = {
					'department': department,
					'profession': profession,
					'reporter_name': reporter_name,
					'total': 0,
					'一般': 0,
					'严重': 0,
					'新的': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			# 统计总数
			reporter_stats[key]['total'] += 1
			
			# 判断是否是新的报告
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			if is_new:
				reporter_stats[key]['新的'] += 1
				# 新的+严重程度组合
				if record.severity == '一般':
					reporter_stats[key]['新的_一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['新的_严重'] += 1
			else:
				# 非新的记录才统计到一般/严重
				if record.severity == '一般':
					reporter_stats[key]['一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['严重'] += 1
		
		# 转换为列表并计算组合统计
		excel_data = []
		department_subtotals = {}
		total_stats = {
			'total': 0,
			'一般': 0,
			'严重': 0,
			'新的_一般': 0,
			'新的_严重': 0
		}
		
		for stats in reporter_stats.values():
			# 直接使用已经正确计算的新的+严重程度组合统计
			excel_data.append({
				'科室': stats['department'],
				'职业': stats['profession'],
				'姓名': stats['reporter_name'],
				'总计': stats['total'],
				'一般': stats['一般'],
				'严重': stats['严重'],
				'新的一般': stats['新的_一般'],
				'新的严重': stats['新的_严重']
			})
			
			# 计算科室小计
			dept = stats['department']
			if dept not in department_subtotals:
				department_subtotals[dept] = {
					'total': 0,
					'一般': 0,
					'严重': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			department_subtotals[dept]['total'] += stats['total']
			department_subtotals[dept]['一般'] += stats['一般']
			department_subtotals[dept]['严重'] += stats['严重']
			department_subtotals[dept]['新的_一般'] += stats['新的_一般']
			department_subtotals[dept]['新的_严重'] += stats['新的_严重']
			
			# 累计总数
			total_stats['total'] += stats['total']
			total_stats['一般'] += stats['一般']
			total_stats['严重'] += stats['严重']
			total_stats['新的_一般'] += stats['新的_一般']
			total_stats['新的_严重'] += stats['新的_严重']
		
		# 按科室、职业、姓名排序
		excel_data.sort(key=lambda x: (x['科室'], x['职业'], x['姓名']))
		
		# 创建完整的表格数据（包含小计和合计）
		final_data = []
		current_dept = None
		
		for row in excel_data:
			# 如果是新科室，先添加上一个科室的小计
			if current_dept and current_dept != row['科室']:
				if current_dept in department_subtotals:
					subtotal = department_subtotals[current_dept]
					final_data.append({
						'科室': '小计',
						'职业': '',
						'姓名': '',
						'总计': subtotal['total'],
						'一般': subtotal['一般'],
						'严重': subtotal['严重'],
						'新的一般': subtotal['新的_一般'],
						'新的严重': subtotal['新的_严重']
					})
			
			final_data.append(row)
			current_dept = row['科室']
		
		# 添加最后一个科室的小计
		if current_dept and current_dept in department_subtotals:
			subtotal = department_subtotals[current_dept]
			final_data.append({
				'科室': '小计',
				'职业': '',
				'姓名': '',
				'总计': subtotal['total'],
				'一般': subtotal['一般'],
				'严重': subtotal['严重'],
				'新的一般': subtotal['新的_一般'],
				'新的严重': subtotal['新的_严重']
			})
		
		# 添加合计行
		final_data.append({
			'科室': '合计',
			'职业': '',
			'姓名': '',
			'总计': total_stats['total'],
			'一般': '',
			'严重': '',
			'新的一般': '',
			'新的严重': ''
		})
		
		# 创建Excel文件
		output = BytesIO()
		df = pd.DataFrame(final_data)
		
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name='上报明细', index=False, header=False, startrow=2)
			
			# 获取工作表进行格式化
			worksheet = writer.sheets['上报明细']
			apply_report_details_template(worksheet)
			
			# 设置列宽
			worksheet.column_dimensions['A'].width = 12  # 科室
			worksheet.column_dimensions['B'].width = 10  # 职业
			worksheet.column_dimensions['C'].width = 12  # 姓名
			worksheet.column_dimensions['D'].width = 8   # 总计
			worksheet.column_dimensions['E'].width = 8   # 一般
			worksheet.column_dimensions['F'].width = 8   # 严重
			worksheet.column_dimensions['G'].width = 10  # 新的一般
			worksheet.column_dimensions['H'].width = 10  # 新的严重
			
			# 创建图表 - 使用matplotlib
			try:
				import matplotlib
				matplotlib.use('Agg')
				import matplotlib.pyplot as plt
				import numpy as np
				
				plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
				plt.rcParams['axes.unicode_minus'] = False
				
				from openpyxl.drawing.image import Image
				
				# 1. 科室上报数量TOP10柱状图
				dept_totals = [(dept, data['total']) for dept, data in department_subtotals.items()]
				dept_totals.sort(key=lambda x: x[1], reverse=True)
				top_depts = dept_totals[:10]
				
				if top_depts:
					fig1, ax1 = plt.subplots(figsize=(10, 6))
					depts = [d[0] for d in top_depts]
					totals = [d[1] for d in top_depts]
					bars = ax1.barh(depts[::-1], totals[::-1], color='#409EFF')
					ax1.set_title('科室上报数量TOP10', fontsize=14, fontweight='bold')
					ax1.set_xlabel('数量')
					for bar, val in zip(bars, totals[::-1]):
						ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2, str(val), va='center', fontsize=10)
					plt.tight_layout()
					
					chart1_image = BytesIO()
					plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
					chart1_image.seek(0)
					plt.close()
					
					img1 = Image(chart1_image)
					img1.width = 500
					img1.height = 300
					worksheet.add_image(img1, 'J2')
				
				# 2. 报告类型占比饼图
				type_data = [
					('一般', total_stats['一般']),
					('严重', total_stats['严重']),
					('新的+一般', total_stats['新的_一般']),
					('新的+严重', total_stats['新的_严重'])
				]
				type_data = [(t, v) for t, v in type_data if v > 0]
				
				if type_data:
					fig2, ax2 = plt.subplots(figsize=(10, 7))
					labels = [f"{t[0]}: {t[1]}例" for t in type_data]
					values = [t[1] for t in type_data]
					colors = ['#67C23A', '#F56C6C', '#409EFF', '#E6A23C']
					draw_pie_with_labels(ax2, values, labels, colors, '报告类型占比')
					
					chart2_image = BytesIO()
					plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
					chart2_image.seek(0)
					plt.close()
					
					img2 = Image(chart2_image)
					img2.width = 450
					img2.height = 350
					worksheet.add_image(img2, 'J20')
					
			except Exception as chart_error:
				print(f"生成上报明细图表时出错: {chart_error}")
		
		output.seek(0)
		
		# 生成文件名
		month_str = get_month_range_str(start_month, end_month)
		filename = f"医院药品不良反应上报明细_{month_str}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=filename
		)
		
	except Exception as e:
		return jsonify({"message": f"导出上报明细失败: {str(e)}"}), 500

@data_bp.route("/analysis/reward-calculation", methods=["GET"])
@require_auth
def get_reward_calculation():
	"""获取奖励计算数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有去重后的记录
		all_records = deduplicated_query.all()
		
		# 按报告人统计
		reporter_stats = {}
		
		for record in all_records:
			# 获取科室信息
			department = "药剂科"
			if record.reporter_profession and "GCP" in str(record.reporter_profession):
				department = "GCP"
			
			profession = record.reporter_profession or "药师"
			reporter_name = record.reporter_signature or "未知"
			
			# 创建唯一键
			key = f"{department}|{profession}|{reporter_name}"
			
			if key not in reporter_stats:
				reporter_stats[key] = {
					'department': department,
					'profession': profession,
					'reporter_name': reporter_name,
					'一般': 0,
					'严重': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			# 判断是否是新的报告
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			if is_new:
				# 新的+严重程度组合
				if record.severity == '一般':
					reporter_stats[key]['新的_一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['新的_严重'] += 1
			else:
				# 非新的记录才统计到一般/严重
				if record.severity == '一般':
					reporter_stats[key]['一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['严重'] += 1
		
		# 计算奖励
		def calculate_tiered_reward(count, base_reward):
			"""阶梯式奖励计算"""
			if count == 0:
				return 0
			
			total_reward = 0
			remaining = count
			
			# 1-5例：100%
			if remaining > 0:
				tier1 = min(remaining, 5)
				total_reward += tier1 * base_reward * 1.0
				remaining -= tier1
			
			# 6-10例：80%
			if remaining > 0:
				tier2 = min(remaining, 5)
				total_reward += tier2 * base_reward * 0.8
				remaining -= tier2
			
			# 11-15例：60%
			if remaining > 0:
				tier3 = min(remaining, 5)
				total_reward += tier3 * base_reward * 0.6
				remaining -= tier3
			
			# 16例以上：40%
			if remaining > 0:
				total_reward += remaining * base_reward * 0.4
			
			return total_reward
		
		# 转换为列表并计算奖励
		reward_details = []
		total_reward = 0

		for stats in reporter_stats.values():
			# 只统计药师和护士张佳丽，其余人员完全剔除（不计入明细、小计、合计）
			if not (stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽'):
				continue
			# 特殊处理：张佳丽固定按药师的奖励规则计算
			if stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽':
				# 药师按照阶梯式计算
				reward_一般 = calculate_tiered_reward(stats['一般'], 50)
				reward_严重 = calculate_tiered_reward(stats['严重'], 70)
				reward_新的一般 = calculate_tiered_reward(stats['新的_一般'], 100)
				reward_新的严重 = calculate_tiered_reward(stats['新的_严重'], 150)

				individual_total = reward_一般 + reward_严重 + reward_新的一般 + reward_新的严重
			else:
				# 非药师按300元/例
				total_cases = stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
				individual_total = total_cases * 300
				reward_一般 = stats['一般'] * 300 if stats['一般'] > 0 else 0
				reward_严重 = stats['严重'] * 300 if stats['严重'] > 0 else 0
				reward_新的一般 = stats['新的_一般'] * 300 if stats['新的_一般'] > 0 else 0
				reward_新的严重 = stats['新的_严重'] * 300 if stats['新的_严重'] > 0 else 0

			reward_detail = {
				'department': stats['department'],
				'profession': stats['profession'],
				'reporter_name': stats['reporter_name'],
				'总数量': stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重'],
				'一般': stats['一般'],
				'奖励_一般': int(reward_一般),
				'严重': stats['严重'],
				'奖励_严重': int(reward_严重),
				'新的_一般': stats['新的_一般'],
				'奖励_新的一般': int(reward_新的一般),
				'新的_严重': stats['新的_严重'],
				'奖励_新的严重': int(reward_新的严重),
				'个人奖励合计': int(individual_total)
			}
			
			reward_details.append(reward_detail)
			total_reward += individual_total
		
		# 按科室、职业、姓名排序
		reward_details.sort(key=lambda x: (x['department'], x['profession'], x['reporter_name']))
		
		# 计算科室小计
		department_subtotals = {}
		for detail in reward_details:
			dept = detail['department']
			if dept not in department_subtotals:
				department_subtotals[dept] = {
					'总数量': 0,
					'一般': 0,
					'奖励_一般': 0,
					'严重': 0,
					'奖励_严重': 0,
					'新的_一般': 0,
					'奖励_新的一般': 0,
					'新的_严重': 0,
					'奖励_新的严重': 0,
					'个人奖励合计': 0
				}

			department_subtotals[dept]['总数量'] += detail['总数量']
			department_subtotals[dept]['一般'] += detail['一般']
			department_subtotals[dept]['奖励_一般'] += detail['奖励_一般']
			department_subtotals[dept]['严重'] += detail['严重']
			department_subtotals[dept]['奖励_严重'] += detail['奖励_严重']
			department_subtotals[dept]['新的_一般'] += detail['新的_一般']
			department_subtotals[dept]['奖励_新的一般'] += detail['奖励_新的一般']
			department_subtotals[dept]['新的_严重'] += detail['新的_严重']
			department_subtotals[dept]['奖励_新的严重'] += detail['奖励_新的严重']
			department_subtotals[dept]['个人奖励合计'] += detail['个人奖励合计']
		
		# 计算总计
		total_stats = {
			'总数量': sum(detail['总数量'] for detail in reward_details),
			'一般': sum(detail['一般'] for detail in reward_details),
			'奖励_一般': sum(detail['奖励_一般'] for detail in reward_details),
			'严重': sum(detail['严重'] for detail in reward_details),
			'奖励_严重': sum(detail['奖励_严重'] for detail in reward_details),
			'新的_一般': sum(detail['新的_一般'] for detail in reward_details),
			'奖励_新的一般': sum(detail['奖励_新的一般'] for detail in reward_details),
			'新的_严重': sum(detail['新的_严重'] for detail in reward_details),
			'奖励_新的严重': sum(detail['奖励_新的严重'] for detail in reward_details),
			'个人奖励合计': int(total_reward)
		}
		
		return jsonify({
			'reward_details': reward_details,
			'department_subtotals': department_subtotals,
			'total_stats': total_stats
		})
		
	except Exception as e:
		return jsonify({"message": f"获取奖励计算数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/export-reward-calculation", methods=["GET"])
@require_auth
def export_reward_calculation():
	"""导出奖励计算数据为Excel"""
	try:
		import pandas as pd
		from io import BytesIO
		from flask import send_file
		
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有去重后的记录
		all_records = deduplicated_query.all()
		
		# 按报告人统计
		reporter_stats = {}
		
		for record in all_records:
			# 获取科室信息
			department = "药剂科"
			if record.reporter_profession and "GCP" in str(record.reporter_profession):
				department = "GCP"
			
			profession = record.reporter_profession or "药师"
			reporter_name = record.reporter_signature or "未知"
			
			# 创建唯一键
			key = f"{department}|{profession}|{reporter_name}"
			
			if key not in reporter_stats:
				reporter_stats[key] = {
					'department': department,
					'profession': profession,
					'reporter_name': reporter_name,
					'一般': 0,
					'严重': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
			
			# 判断是否是新的报告
			is_new = (record.report_type_new and 
					 record.report_type_new.strip() and 
					 record.report_type_new == '新的')
			
			if is_new:
				# 新的+严重程度组合
				if record.severity == '一般':
					reporter_stats[key]['新的_一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['新的_严重'] += 1
			else:
				# 非新的记录才统计到一般/严重
				if record.severity == '一般':
					reporter_stats[key]['一般'] += 1
				elif record.severity == '严重':
					reporter_stats[key]['严重'] += 1
		
		# 计算奖励
		def calculate_tiered_reward(count, base_reward):
			"""阶梯式奖励计算"""
			if count == 0:
				return 0
			
			total_reward = 0
			remaining = count
			
			# 1-5例：100%
			if remaining > 0:
				tier1 = min(remaining, 5)
				total_reward += tier1 * base_reward * 1.0
				remaining -= tier1
			
			# 6-10例：80%
			if remaining > 0:
				tier2 = min(remaining, 5)
				total_reward += tier2 * base_reward * 0.8
				remaining -= tier2
			
			# 11-15例：60%
			if remaining > 0:
				tier3 = min(remaining, 5)
				total_reward += tier3 * base_reward * 0.6
				remaining -= tier3
			
			# 16例以上：40%
			if remaining > 0:
				total_reward += remaining * base_reward * 0.4
			
			return total_reward
		
		# 创建Excel数据
		excel_data = []
		department_subtotals = {}
		total_stats = {
			'总数量': 0,
			'一般': 0,
			'奖励_一般': 0,
			'严重': 0,
			'奖励_严重': 0,
			'新的_一般': 0,
			'奖励_新的一般': 0,
			'新的_严重': 0,
			'奖励_新的严重': 0,
			'个人奖励合计': 0
		}

		for stats in reporter_stats.values():
			# 只统计药师和护士张佳丽，其余人员完全剔除（不计入明细、小计、合计）
			if not (stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽'):
				continue
			# 特殊处理：张佳丽固定按药师的奖励规则计算
			if stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽':
				# 药师按照阶梯式计算
				reward_一般 = calculate_tiered_reward(stats['一般'], 50)
				reward_严重 = calculate_tiered_reward(stats['严重'], 70)
				reward_新的一般 = calculate_tiered_reward(stats['新的_一般'], 100)
				reward_新的严重 = calculate_tiered_reward(stats['新的_严重'], 150)

				individual_total = reward_一般 + reward_严重 + reward_新的一般 + reward_新的严重
			else:
				# 非药师按300元/例
				total_cases = stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
				individual_total = total_cases * 300
				reward_一般 = stats['一般'] * 300 if stats['一般'] > 0 else 0
				reward_严重 = stats['严重'] * 300 if stats['严重'] > 0 else 0
				reward_新的一般 = stats['新的_一般'] * 300 if stats['新的_一般'] > 0 else 0
				reward_新的严重 = stats['新的_严重'] * 300 if stats['新的_严重'] > 0 else 0

			excel_data.append({
				'科室': stats['department'],
				'职业': stats['profession'],
				'姓名': stats['reporter_name'],
				'总数量': stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重'],
				'一般': stats['一般'],
				'奖励/元': int(reward_一般),
				'严重': stats['严重'],
				'奖励/元.1': int(reward_严重),
				'新的一般': stats['新的_一般'],
				'奖励/元.2': int(reward_新的一般),
				'新的严重': stats['新的_严重'],
				'奖励/元.3': int(reward_新的严重),
				'个人奖励合计/元': int(individual_total)
			})
			
			# 计算科室小计
			dept = stats['department']
			if dept not in department_subtotals:
				department_subtotals[dept] = {
					'总数量': 0,
					'一般': 0,
					'奖励_一般': 0,
					'严重': 0,
					'奖励_严重': 0,
					'新的_一般': 0,
					'奖励_新的一般': 0,
					'新的_严重': 0,
					'奖励_新的严重': 0,
					'个人奖励合计': 0
				}

			department_subtotals[dept]['总数量'] += stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
			department_subtotals[dept]['一般'] += stats['一般']
			department_subtotals[dept]['奖励_一般'] += int(reward_一般)
			department_subtotals[dept]['严重'] += stats['严重']
			department_subtotals[dept]['奖励_严重'] += int(reward_严重)
			department_subtotals[dept]['新的_一般'] += stats['新的_一般']
			department_subtotals[dept]['奖励_新的一般'] += int(reward_新的一般)
			department_subtotals[dept]['新的_严重'] += stats['新的_严重']
			department_subtotals[dept]['奖励_新的严重'] += int(reward_新的严重)
			department_subtotals[dept]['个人奖励合计'] += int(individual_total)
			
			# 累计总数
			total_stats['总数量'] += stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
			total_stats['一般'] += stats['一般']
			total_stats['奖励_一般'] += int(reward_一般)
			total_stats['严重'] += stats['严重']
			total_stats['奖励_严重'] += int(reward_严重)
			total_stats['新的_一般'] += stats['新的_一般']
			total_stats['奖励_新的一般'] += int(reward_新的一般)
			total_stats['新的_严重'] += stats['新的_严重']
			total_stats['奖励_新的严重'] += int(reward_新的严重)
			total_stats['个人奖励合计'] += int(individual_total)
		
		# 按科室、职业、姓名排序
		excel_data.sort(key=lambda x: (x['科室'], x['职业'], x['姓名']))
		
		# 创建完整的表格数据（包含小计和合计）
		final_data = []
		current_dept = None
		
		for row in excel_data:
			# 如果是新科室，先添加上一个科室的小计
			if current_dept and current_dept != row['科室']:
				if current_dept in department_subtotals:
					subtotal = department_subtotals[current_dept]
					final_data.append({
						'科室': '小计',
						'职业': '',
						'姓名': '',
						'总数量': subtotal['一般'] + subtotal['严重'] + subtotal['新的_一般'] + subtotal['新的_严重'],
						'一般': subtotal['一般'],
						'奖励/元': subtotal['奖励_一般'],
						'严重': subtotal['严重'],
						'奖励/元.1': subtotal['奖励_严重'],
						'新的一般': subtotal['新的_一般'],
						'奖励/元.2': subtotal['奖励_新的一般'],
						'新的严重': subtotal['新的_严重'],
						'奖励/元.3': subtotal['奖励_新的严重'],
						'个人奖励合计/元': subtotal['个人奖励合计']
					})
			
			final_data.append(row)
			current_dept = row['科室']
		
		# 添加最后一个科室的小计
		if current_dept and current_dept in department_subtotals:
			subtotal = department_subtotals[current_dept]
			final_data.append({
				'科室': '小计',
				'职业': '',
				'姓名': '',
				'总数量': subtotal['一般'] + subtotal['严重'] + subtotal['新的_一般'] + subtotal['新的_严重'],
				'一般': subtotal['一般'],
				'奖励/元': subtotal['奖励_一般'],
				'严重': subtotal['严重'],
				'奖励/元.1': subtotal['奖励_严重'],
				'新的一般': subtotal['新的_一般'],
				'奖励/元.2': subtotal['奖励_新的一般'],
				'新的严重': subtotal['新的_严重'],
				'奖励/元.3': subtotal['奖励_新的严重'],
				'个人奖励合计/元': subtotal['个人奖励合计']
			})

		# 添加合计行
		final_data.append({
			'科室': '合计',
			'职业': '',
			'姓名': '',
			'总数量': '',
			'一般': '',
			'奖励/元': '',
			'严重': '',
			'奖励/元.1': '',
			'新的一般': '',
			'奖励/元.2': '',
			'新的严重': '',
			'奖励/元.3': '',
			'个人奖励合计/元': total_stats['个人奖励合计']
		})
		
		# 创建Excel文件
		output = BytesIO()
		df = pd.DataFrame(final_data)
		
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			df.to_excel(writer, sheet_name='奖励计算', index=False, header=False, startrow=2)
			
			# 获取工作表进行格式化
			worksheet = writer.sheets['奖励计算']
			apply_reward_template(worksheet)
			
			# 设置列宽
			worksheet.column_dimensions['A'].width = 12  # 科室
			worksheet.column_dimensions['B'].width = 10  # 职业
			worksheet.column_dimensions['C'].width = 12  # 姓名
			worksheet.column_dimensions['D'].width = 10  # 总数量
			worksheet.column_dimensions['E'].width = 8   # 一般
			worksheet.column_dimensions['F'].width = 10  # 奖励/元
			worksheet.column_dimensions['G'].width = 8   # 严重
			worksheet.column_dimensions['H'].width = 10  # 奖励/元.1
			worksheet.column_dimensions['I'].width = 10  # 新的一般
			worksheet.column_dimensions['J'].width = 10  # 奖励/元.2
			worksheet.column_dimensions['K'].width = 10  # 新的严重
			worksheet.column_dimensions['L'].width = 10  # 奖励/元.3
			worksheet.column_dimensions['M'].width = 15  # 个人奖励合计/元
			
			# 创建图表 - 使用matplotlib
			try:
				import matplotlib
				matplotlib.use('Agg')
				import matplotlib.pyplot as plt
				
				plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
				plt.rcParams['axes.unicode_minus'] = False
				
				from openpyxl.drawing.image import Image
				
				# 准备个人奖励数据（排除小计和合计行）
				person_rewards = [(row['姓名'], row['个人奖励合计/元']) 
								  for row in excel_data if row['姓名'] and row['科室'] not in ['小计', '合计']]
				person_rewards.sort(key=lambda x: x[1], reverse=True)
				top_persons = person_rewards[:10]
				
				# 1. 个人奖励TOP10柱状图
				if top_persons:
					fig1, ax1 = plt.subplots(figsize=(10, 6))
					names = [p[0] for p in top_persons]
					rewards = [p[1] for p in top_persons]
					bars = ax1.barh(names[::-1], rewards[::-1], color='#E6A23C')
					ax1.set_title('个人奖励TOP10', fontsize=14, fontweight='bold')
					ax1.set_xlabel('奖励金额(元)')
					for bar, val in zip(bars, rewards[::-1]):
						ax1.text(bar.get_width() + 10, bar.get_y() + bar.get_height()/2, f'¥{val}', va='center', fontsize=10)
					plt.tight_layout()
					
					chart1_image = BytesIO()
					plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
					chart1_image.seek(0)
					plt.close()
					
					img1 = Image(chart1_image)
					img1.width = 500
					img1.height = 300
					worksheet.add_image(img1, 'O2')
				
				# 2. 奖励类型占比饼图
				type_rewards = [
					('一般', total_stats['奖励_一般']),
					('严重', total_stats['奖励_严重']),
					('新的+一般', total_stats['奖励_新的一般']),
					('新的+严重', total_stats['奖励_新的严重'])
				]
				type_rewards = [(t, v) for t, v in type_rewards if v > 0]
				
				if type_rewards:
					fig2, ax2 = plt.subplots(figsize=(10, 7))
					labels = [f"{t[0]}: ¥{t[1]}" for t in type_rewards]
					values = [t[1] for t in type_rewards]
					colors = ['#67C23A', '#F56C6C', '#409EFF', '#E6A23C']
					draw_pie_with_labels(ax2, values, labels, colors, '奖励类型占比')
					
					chart2_image = BytesIO()
					plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
					chart2_image.seek(0)
					plt.close()
					
					img2 = Image(chart2_image)
					img2.width = 450
					img2.height = 350
					worksheet.add_image(img2, 'O20')
					
			except Exception as chart_error:
				print(f"生成奖励计算图表时出错: {chart_error}")
		
		output.seek(0)
		
		# 生成文件名
		month_str = get_month_range_str(start_month, end_month)
		filename = f"药品不良反应奖励计算_{month_str}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=filename
		)
		
	except Exception as e:
		return jsonify({"message": f"导出奖励计算数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/drug-summary", methods=["GET"])
@require_auth
def get_drug_summary():
	"""获取药品汇总数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有去重后的记录
		all_records = deduplicated_query.all()
		
		# 按药品统计
		drug_stats = {}
		
		for record in all_records:
			# 获取药品信息
			drug_name = record.generic_name or "未知药品"
			manufacturer = record.manufacturer or "未知厂家"
			
			# 创建唯一键
			key = f"{drug_name}|{manufacturer}"
			
			if key not in drug_stats:
				drug_stats[key] = {
					'drug_name': drug_name,
					'manufacturer': manufacturer,
					'一般': 0,
					'严重': 0,
					'total': 0
				}
			
			# 统计严重程度
			if record.severity == '一般':
				drug_stats[key]['一般'] += 1
			elif record.severity == '严重':
				drug_stats[key]['严重'] += 1
			
			drug_stats[key]['total'] += 1
		
		# 转换为列表并按合计数量排序
		drug_summary = list(drug_stats.values())
		drug_summary.sort(key=lambda x: x['total'], reverse=True)
		
		return jsonify({
			'drug_summary': drug_summary,
			'reaction_summary': build_reaction_summary(all_records)
		})
		
	except Exception as e:
		return jsonify({"message": f"获取药品汇总数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/export-drug-summary", methods=["GET"])
@require_auth
def export_drug_summary():
	"""导出药品汇总数据为Excel，包含前十名柱状图和所有合计的饼状图"""
	try:
		import pandas as pd
		from io import BytesIO
		from flask import send_file
		
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		# 获取所有去重后的记录
		all_records = get_deduplicated_records(start_month, end_month).all()
		
		# 按药品统计
		drug_stats = {}
		
		for record in all_records:
			# 获取药品信息
			drug_name = record.generic_name or "未知药品"
			manufacturer = record.manufacturer or "未知厂家"
			
			# 创建唯一键
			key = f"{drug_name}|{manufacturer}"
			
			if key not in drug_stats:
				drug_stats[key] = {
					'drug_name': drug_name,
					'manufacturer': manufacturer,
					'一般': 0,
					'严重': 0,
					'total': 0
				}
			
			# 统计严重程度
			if record.severity == '一般':
				drug_stats[key]['一般'] += 1
			elif record.severity == '严重':
				drug_stats[key]['严重'] += 1
			
			drug_stats[key]['total'] += 1
		
		# 转换为列表并按合计数量排序
		drug_list = list(drug_stats.values())
		drug_list.sort(key=lambda x: x['total'], reverse=True)
		
		# 创建Excel数据
		excel_data = []
		for i, drug in enumerate(drug_list, 1):
			excel_data.append({
				'序号': i,
				'通用名称': drug['drug_name'],
				'生产厂家': drug['manufacturer'],
				'严重': drug['严重'],
				'一般': drug['一般'],
				'合计': drug['total']
			})
		
		# 添加合计行
		if excel_data:
			total_severe = sum(item['严重'] for item in excel_data)
			total_general = sum(item['一般'] for item in excel_data)
			excel_data.append({
				'序号': '',
				'通用名称': '合计',
				'生产厂家': '',
				'严重': total_severe,
				'一般': total_general,
				'合计': total_severe + total_general
			})
		
		# 创建Excel文件
		output = BytesIO()
		
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			# 创建数据表
			df = pd.DataFrame(excel_data)
			df.to_excel(writer, sheet_name='发生不良反应的药品汇总', index=False)
			df_reaction = pd.DataFrame(build_reaction_summary_excel_data(all_records))
			df_reaction.to_excel(writer, sheet_name='不良反应汇总分析', index=False)
			add_reaction_summary_charts(writer, all_records)
			
			# 创建图表 - 使用matplotlib
			try:
				import matplotlib
				matplotlib.use('Agg')
				import matplotlib.pyplot as plt
				import numpy as np
				
				plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
				plt.rcParams['axes.unicode_minus'] = False
				
				from openpyxl.drawing.image import Image
				worksheet = writer.sheets['发生不良反应的药品汇总']
				
				# 1. 前十名药品柱状图
				top_10_drugs = drug_list[:10]
				if top_10_drugs:
					fig1, ax1 = plt.subplots(figsize=(12, 7))
					drug_names = [d['drug_name'][:15] + '..' if len(d['drug_name']) > 15 else d['drug_name'] for d in top_10_drugs]
					general_counts = [d['一般'] for d in top_10_drugs]
					severe_counts = [d['严重'] for d in top_10_drugs]
					
					x = np.arange(len(drug_names))
					width = 0.35
					bars1 = ax1.barh(x - width/2, general_counts[::-1], width, label='一般', color='#67C23A')
					bars2 = ax1.barh(x + width/2, severe_counts[::-1], width, label='严重', color='#F56C6C')
					
					ax1.set_yticks(x)
					ax1.set_yticklabels(drug_names[::-1], fontsize=9)
					ax1.set_title('药品不良反应前十名', fontsize=14, fontweight='bold')
					ax1.set_xlabel('数量')
					ax1.legend()
					plt.tight_layout()
					
					chart1_image = BytesIO()
					plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
					chart1_image.seek(0)
					plt.close()
					
					img1 = Image(chart1_image)
					img1.width = 600
					img1.height = 350
					worksheet.add_image(img1, 'H2')
				
				# 2. 药品分布饼图
				top_10_for_pie = drug_list[:10]
				if top_10_for_pie:
					fig2, ax2 = plt.subplots(figsize=(12, 9))
					labels = [f"{d['drug_name']}: {d['total']}例" for d in top_10_for_pie]
					values = [d['total'] for d in top_10_for_pie]
					other_total = sum(d['total'] for d in drug_list[10:])
					if other_total > 0:
						labels.append(f'其他: {other_total}例')
						values.append(other_total)
					
					colors = ['#409EFF', '#67C23A', '#E6A23C', '#F56C6C', '#909399', '#606266', '#E91E63', '#9C27B0', '#3F51B5', '#00BCD4', '#CCCCCC']
					draw_pie_with_labels(ax2, values, labels, colors, '药品不良反应分布')
					
					chart2_image = BytesIO()
					plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
					chart2_image.seek(0)
					plt.close()
					
					img2 = Image(chart2_image)
					img2.width = 600
					img2.height = 450
					worksheet.add_image(img2, 'H24')
				
			except Exception as chart_error:
				print(f"生成图表时出错: {chart_error}")
		
		output.seek(0)
		
		# 生成文件名
		filename = f"发生不良反应的药品汇总_{get_month_range_str(start_month, end_month)}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=filename
		)
		
	except Exception as e:
		return jsonify({"message": f"导出药品汇总数据失败: {str(e)}"}), 500

@data_bp.route("/analysis/export-all-tabs", methods=["GET"])
@require_auth
def export_all_tabs():
	"""导出所有Tab数据到一个Excel文件的多个Sheet - 复用现有导出功能"""
	try:
		import pandas as pd
		from io import BytesIO
		from flask import send_file
		from datetime import datetime
		
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		# 创建最终的Excel文件
		output = BytesIO()

		with pd.ExcelWriter(output, engine='openpyxl') as writer:

			# 1. 全部数据
			query_all = AdverseReactionReport.query
			query_all = apply_month_filter(query_all, start_month, end_month)
			
			all_reports = query_all.order_by(AdverseReactionReport.created_at.desc()).all()
			
			# 转换为DataFrame
			all_data = []
			for report in all_reports:
				all_data.append({
					'报告表编码': report.report_code,
					'报告类型-新的': report.report_type_new or '',
					'报告类型-严重程度': report.severity,
					'病历号/门诊号': report.medical_record_no,
					'怀疑/并用': report.suspect_concurrent,
					'通用名称': report.generic_name,
					'生产厂家': report.manufacturer,
					'不良反应名称': report.adverse_reaction_name,
					'报告人职业': report.reporter_profession,
					'报告人签名': report.reporter_signature,
					'国家中心接收时间': report.national_center_receive_time.strftime('%Y-%m-%d %H:%M:%S') if report.national_center_receive_time else ''
				})
			
			df_all = pd.DataFrame(all_data)
			df_all.to_excel(writer, sheet_name='全部数据', index=False)
			
			# 2. 去重数据
			query_dedup = AdverseReactionReport.query
			query_dedup = apply_month_filter(query_dedup, start_month, end_month)
			
			# 先过滤掉"并用"的数据，只保留"怀疑"的数据
			filtered_query = query_dedup.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
			
			# 然后按报告编码去重
			subquery = filtered_query.with_entities(
				AdverseReactionReport.report_code,
				db.func.min(AdverseReactionReport.id).label('min_id')
			).group_by(AdverseReactionReport.report_code).subquery()
			
			dedup_reports = db.session.query(AdverseReactionReport).join(
				subquery, AdverseReactionReport.id == subquery.c.min_id
			).order_by(AdverseReactionReport.created_at.desc()).all()
			
			# 转换为DataFrame
			dedup_data = []
			for report in dedup_reports:
				dedup_data.append({
					'报告表编码': report.report_code,
					'报告类型-新的': report.report_type_new or '',
					'报告类型-严重程度': report.severity,
					'病历号/门诊号': report.medical_record_no,
					'怀疑/并用': report.suspect_concurrent,
					'通用名称': report.generic_name,
					'生产厂家': report.manufacturer,
					'不良反应名称': report.adverse_reaction_name,
					'报告人职业': report.reporter_profession,
					'报告人签名': report.reporter_signature,
					'国家中心接收时间': report.national_center_receive_time.strftime('%Y-%m-%d %H:%M:%S') if report.national_center_receive_time else ''
				})
			
			df_dedup = pd.DataFrame(dedup_data)
			df_dedup.to_excel(writer, sheet_name='去重数据', index=False)
			
			# 3. 类型统计数据 - 复用export_statistics的核心逻辑
			try:
				# 使用去重数据进行统计
				all_records = dedup_reports
				
				# 正确的统计逻辑：避免重复计算
				new_count = 0  # "报告类型-新的"为"新的"的数量
				severity_counts = {'严重': 0, '一般': 0}  # 按严重程度统计（所有记录）
				non_new_severity_counts = {'严重': 0, '一般': 0}  # 非新的按严重程度统计
				new_and_severity_counts = {'严重': 0, '一般': 0}  # 既是新的又是对应严重程度的记录数量
				
				for record in all_records:
					# 判断是否为"新的"
					is_new = (record.report_type_new and 
							 record.report_type_new.strip() and 
							 record.report_type_new == '新的')
					
					# 统计"报告类型-新的"为"新的"的记录
					if is_new:
						new_count += 1
						# 统计既是新的又是对应严重程度的记录
						if record.severity in new_and_severity_counts:
							new_and_severity_counts[record.severity] += 1
					else:
						# 统计非新的按严重程度
						if record.severity in non_new_severity_counts:
							non_new_severity_counts[record.severity] += 1
					
					# 统计所有记录的严重程度
					if record.severity in severity_counts:
						severity_counts[record.severity] += 1
				
				# 组合统计结果（避免重复计算）
				type_counts = {
					'新的+严重': new_count + severity_counts['严重'] - new_and_severity_counts['严重'],
					'新的+一般': new_count + severity_counts['一般'] - new_and_severity_counts['一般'],
					'严重': non_new_severity_counts['严重'],
					'一般': non_new_severity_counts['一般']
				}
				
				# 创建类型统计数据
				display_order = ['一般', '严重', '新的+一般', '新的+严重']
				severity_data = []
				total_count = len(all_records)  # 总计是去重之后的总数
				
				for type_name in display_order:
					if type_name in type_counts:
						count = type_counts[type_name]
						percentage = (count / total_count * 100) if total_count > 0 else 0
						severity_data.append({
							'报告类型': type_name,
							'数量': count,
							'占比(%)': round(percentage, 1)
						})
				
				# 添加总计行
				severity_data.append({
					'报告类型': '总计',
					'数量': total_count,
					'占比(%)': 100.0
				})
				
				df_statistics = pd.DataFrame(severity_data)
				df_statistics.to_excel(writer, sheet_name='类型统计数据', index=False, startrow=0)
				
				# 添加ADR数量占比统计数据
				# 按月分组统计去重数据
				monthly_all_data = {}
				for record in all_records:
					month_key = record.national_center_receive_time.strftime('%Y%m')
					
					if month_key not in monthly_all_data:
						monthly_all_data[month_key] = {
							'total': 0,
							'严重的': 0,
							'新的': 0
						}
					
					monthly_all_data[month_key]['total'] += 1
					
					# 统计严重的
					if record.severity == '严重':
						monthly_all_data[month_key]['严重的'] += 1
					
					# 统计新的
					if (record.report_type_new and 
						record.report_type_new.strip() and 
						record.report_type_new == '新的'):
						monthly_all_data[month_key]['新的'] += 1
				
				# 生成12个月份的数据
				months = []
				for i in range(1, 13):
					month_key = f"2025{i:02d}"
					months.append(month_key)
				
				# 创建表格数据
				total_row = ['总计'] + [monthly_all_data.get(m, {}).get('total', 0) for m in months]
				serious_row = ['严重的'] + [monthly_all_data.get(m, {}).get('严重的', 0) for m in months]
				serious_ratio_row = ['占比'] + []
				serious_plus_new_row = ['严重的+新的'] + []
				serious_plus_new_ratio_row = ['占比'] + []
				
				# 计算占比和严重的+新的
				for month_key in months:
					data = monthly_all_data.get(month_key, {'total': 0, '严重的': 0, '新的': 0})
					total = data['total']
					serious_count = data['严重的']
					new_count = data['新的']
					serious_plus_new_count = serious_count + new_count
					
					# 计算占比
					serious_ratio = (serious_count / total * 100) if total > 0 else 0
					serious_plus_new_ratio = (serious_plus_new_count / total * 100) if total > 0 else 0
					
					serious_ratio_row.append(f"{serious_ratio:.2f}%")
					serious_plus_new_row.append(serious_plus_new_count)
					serious_plus_new_ratio_row.append(f"{serious_plus_new_ratio:.2f}%")
				
				# 空行分隔
				start_row_monthly = len(df_statistics) + 3
				
				# ADR数量占比统计表
				columns = ['ADR数量'] + months
				monthly_data = [
					total_row,
					serious_row,
					serious_ratio_row,
					serious_plus_new_row,
					serious_plus_new_ratio_row
				]
				df_monthly = pd.DataFrame(monthly_data, columns=columns)
				df_monthly.to_excel(writer, sheet_name='类型统计数据', index=False, startrow=start_row_monthly)
				
				# 添加类型分布图表到Excel中
				fig = None
				chart_image = None
				try:
					import matplotlib
					matplotlib.use('Agg')  # 使用非交互式后端
					import matplotlib.pyplot as plt
					from matplotlib import font_manager
					import numpy as np
					
					# 设置中文字体，如果没有中文字体就使用默认字体
					try:
						plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
						plt.rcParams['axes.unicode_minus'] = False
					except:
						# 如果字体设置失败，使用默认字体
						pass
					
					# 准备图表数据，只显示总计、严重、新的+严重
					chart_data_filtered = []
					
					# 添加总计
					chart_data_filtered.append({
						'name': '总计',
						'value': total_count
					})
					
					# 只添加严重和新的+严重
					for item in severity_data[:-1]:  # 排除最后的总计行
						if item['报告类型'] in ['严重', '新的+严重']:
							chart_data_filtered.append({
								'name': item['报告类型'],
								'value': item['数量']
							})
					
					chart_types = [item['name'] for item in chart_data_filtered]
					chart_values = [item['value'] for item in chart_data_filtered]
					chart_colors = ['#606266', '#F56C6C', '#E6A23C']  # 总计、严重、新的+严重
					
					# 创建图表
					fig, ax = plt.subplots(figsize=(10, 6))
					bars = ax.bar(chart_types, chart_values, color=chart_colors[:len(chart_types)], width=0.5)
					
					# 设置图表样式
					ax.set_title('类型分布图', fontsize=16, fontweight='bold', pad=20)
					ax.set_ylabel('数量', fontsize=12)
					ax.set_xlabel('报告类型', fontsize=12)
					
					# 在柱子上显示数值
					for bar, value in zip(bars, chart_values):
						height = bar.get_height()
						ax.text(bar.get_x() + bar.get_width()/2., height + 0.5,
							   f'{value}', ha='center', va='bottom', fontsize=11, fontweight='bold')
					
					# 设置y轴范围，确保有足够空间显示数值
					max_value = max(chart_values) if chart_values else 1
					ax.set_ylim(0, max_value * 1.15)
					
					# 美化图表
					ax.grid(True, alpha=0.3, axis='y')
					ax.set_axisbelow(True)
					
					# 调整布局
					plt.tight_layout()
					
					# 保存图表为内存中的图片
					chart_image = BytesIO()
					plt.savefig(chart_image, format='png', dpi=300, bbox_inches='tight')
					chart_image.seek(0)
					
					# 读取图片数据到内存中，避免I/O问题
					image_data = chart_image.getvalue()
					image_buffer = BytesIO(image_data)
					
					# 获取工作表并插入图片
					worksheet = writer.sheets['类型统计数据']
					
					# 插入图片到右侧空白区域
					from openpyxl.drawing.image import Image
					img = Image(image_buffer)
					
					# 调整图片大小
					img.width = 600
					img.height = 360
					
					# 插入图片到合适位置（右侧）
					worksheet.add_image(img, 'H2')
					
				except Exception as chart_error:
					print(f"生成类型分布图时出错: {chart_error}")
					# 如果图表生成失败，继续执行其他部分
				finally:
					# 确保资源被正确清理
					try:
						if fig is not None:
							plt.close(fig)
						if chart_image is not None:
							chart_image.close()
						# 清理所有matplotlib图形
						plt.close('all')
					except:
						pass
				
			except Exception as e:
				print(f"类型统计导出错误: {str(e)}")
			
			# 4. 上报明细-全院数据 - 复用export_report_details的核心逻辑
			try:
				# 按报告人统计
				reporter_stats = {}
				
				for record in all_records:
					department = "药剂科"
					if record.reporter_profession and "GCP" in str(record.reporter_profession):
						department = "GCP"
					
					profession = record.reporter_profession or "药师"
					reporter_name = record.reporter_signature or "未知"
					
					key = f"{department}|{profession}|{reporter_name}"
					
					if key not in reporter_stats:
						reporter_stats[key] = {
							'department': department,
							'profession': profession,
							'reporter_name': reporter_name,
							'total': 0,
							'一般': 0,
							'严重': 0,
							'新的_一般': 0,
							'新的_严重': 0
						}
					
					reporter_stats[key]['total'] += 1
					
					is_new = (record.report_type_new and 
							 record.report_type_new.strip() and 
							 record.report_type_new == '新的')
					
					if is_new:
						if record.severity == '一般':
							reporter_stats[key]['新的_一般'] += 1
						elif record.severity == '严重':
							reporter_stats[key]['新的_严重'] += 1
					else:
						if record.severity == '一般':
							reporter_stats[key]['一般'] += 1
						elif record.severity == '严重':
							reporter_stats[key]['严重'] += 1
				
				# 创建上报明细Excel数据
				excel_data = []
				department_subtotals = {}
				total_stats = {
					'total': 0,
					'一般': 0,
					'严重': 0,
					'新的_一般': 0,
					'新的_严重': 0
				}
				
				for stats in reporter_stats.values():
					excel_data.append({
						'科室': stats['department'],
						'职业': stats['profession'],
						'姓名': stats['reporter_name'],
						'总计': stats['total'],
						'一般': stats['一般'],
						'严重': stats['严重'],
						'新的一般': stats['新的_一般'],
						'新的严重': stats['新的_严重']
					})
					
					# 计算科室小计
					dept = stats['department']
					if dept not in department_subtotals:
						department_subtotals[dept] = {
							'total': 0,
							'一般': 0,
							'严重': 0,
							'新的_一般': 0,
							'新的_严重': 0
						}
					
					department_subtotals[dept]['total'] += stats['total']
					department_subtotals[dept]['一般'] += stats['一般']
					department_subtotals[dept]['严重'] += stats['严重']
					department_subtotals[dept]['新的_一般'] += stats['新的_一般']
					department_subtotals[dept]['新的_严重'] += stats['新的_严重']
					
					# 累计总数
					total_stats['total'] += stats['total']
					total_stats['一般'] += stats['一般']
					total_stats['严重'] += stats['严重']
					total_stats['新的_一般'] += stats['新的_一般']
					total_stats['新的_严重'] += stats['新的_严重']
				
				# 按科室、职业、姓名排序
				excel_data.sort(key=lambda x: (x['科室'], x['职业'], x['姓名']))
				
				# 创建完整的表格数据（包含小计和合计）
				final_data = []
				current_dept = None
				
				for row in excel_data:
					# 如果是新科室，先添加上一个科室的小计
					if current_dept and current_dept != row['科室']:
						if current_dept in department_subtotals:
							subtotal = department_subtotals[current_dept]
							final_data.append({
								'科室': '小计',
								'职业': '',
								'姓名': '',
								'总计': subtotal['total'],
								'一般': subtotal['一般'],
								'严重': subtotal['严重'],
								'新的一般': subtotal['新的_一般'],
								'新的严重': subtotal['新的_严重']
							})
					
					final_data.append(row)
					current_dept = row['科室']
				
				# 添加最后一个科室的小计
				if current_dept and current_dept in department_subtotals:
					subtotal = department_subtotals[current_dept]
					final_data.append({
						'科室': '小计',
						'职业': '',
						'姓名': '',
						'总计': subtotal['total'],
						'一般': subtotal['一般'],
						'严重': subtotal['严重'],
						'新的一般': subtotal['新的_一般'],
						'新的严重': subtotal['新的_严重']
					})
				
				# 添加合计行
				final_data.append({
					'科室': '合计',
					'职业': '',
					'姓名': '',
					'总计': total_stats['total'],
					'一般': '',
					'严重': '',
					'新的一般': '',
					'新的严重': ''
				})
				
				df_report_details = pd.DataFrame(final_data)
				df_report_details.to_excel(writer, sheet_name='上报明细-全院', index=False, header=False, startrow=2)
				worksheet_detail = writer.sheets['上报明细-全院']
				apply_report_details_template(worksheet_detail)
				
				# 添加上报明细图表 - 使用matplotlib
				try:
					import matplotlib
					matplotlib.use('Agg')
					import matplotlib.pyplot as plt
					
					plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
					plt.rcParams['axes.unicode_minus'] = False
					
					from openpyxl.drawing.image import Image
					
					# 1. 科室上报数量TOP10柱状图
					dept_totals = [(dept, data['total']) for dept, data in department_subtotals.items()]
					dept_totals.sort(key=lambda x: x[1], reverse=True)
					top_depts = dept_totals[:10]
					
					if top_depts:
						fig1, ax1 = plt.subplots(figsize=(10, 6))
						depts = [d[0] for d in top_depts]
						totals = [d[1] for d in top_depts]
						bars = ax1.barh(depts[::-1], totals[::-1], color='#409EFF')
						ax1.set_title('科室上报数量TOP10', fontsize=14, fontweight='bold')
						ax1.set_xlabel('数量')
						for bar, val in zip(bars, totals[::-1]):
							ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2, str(val), va='center', fontsize=10)
						plt.tight_layout()
						
						chart1_image = BytesIO()
						plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
						chart1_image.seek(0)
						plt.close()
						
						img1 = Image(chart1_image)
						img1.width = 500
						img1.height = 300
						worksheet_detail.add_image(img1, 'J2')
					
					# 2. 报告类型占比饼图
					type_data = [
						('一般', total_stats['一般']),
						('严重', total_stats['严重']),
						('新的+一般', total_stats['新的_一般']),
						('新的+严重', total_stats['新的_严重'])
					]
					type_data = [(t, v) for t, v in type_data if v > 0]
					
					if type_data:
						fig2, ax2 = plt.subplots(figsize=(10, 7))
						labels = [f"{t[0]}: {t[1]}例" for t in type_data]
						values = [t[1] for t in type_data]
						colors = ['#67C23A', '#F56C6C', '#409EFF', '#E6A23C']
						draw_pie_with_labels(ax2, values, labels, colors, '报告类型占比')
						
						chart2_image = BytesIO()
						plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
						chart2_image.seek(0)
						plt.close()
						
						img2 = Image(chart2_image)
						img2.width = 450
						img2.height = 350
						worksheet_detail.add_image(img2, 'J20')
						
				except Exception as chart_err:
					print(f"生成上报明细图表时出错: {chart_err}")
				
			except Exception as e:
				print(f"上报明细导出错误: {str(e)}")
			
			# 5. 奖励计算-药学数据 - 复用export_reward_calculation的核心逻辑
			try:
				def calculate_tiered_reward(count, base_reward):
					if count == 0:
						return 0
					
					total_reward = 0
					remaining = count
					
					# 1-5例：100%
					if remaining > 0:
						tier1 = min(remaining, 5)
						total_reward += tier1 * base_reward * 1.0
						remaining -= tier1
					
					# 6-10例：80%
					if remaining > 0:
						tier2 = min(remaining, 5)
						total_reward += tier2 * base_reward * 0.8
						remaining -= tier2
					
					# 11-15例：60%
					if remaining > 0:
						tier3 = min(remaining, 5)
						total_reward += tier3 * base_reward * 0.6
						remaining -= tier3
					
					# 16例以上：40%
					if remaining > 0:
						total_reward += remaining * base_reward * 0.4
					
					return total_reward
				
				# 创建奖励计算Excel数据
				reward_excel_data = []
				reward_department_subtotals = {}
				reward_total_stats = {
					'总数量': 0,
					'一般': 0,
					'奖励_一般': 0,
					'严重': 0,
					'奖励_严重': 0,
					'新的_一般': 0,
					'奖励_新的一般': 0,
					'新的_严重': 0,
					'奖励_新的严重': 0,
					'个人奖励合计': 0
				}

				for stats in reporter_stats.values():
					# 只统计药师和护士张佳丽，其余人员完全剔除（不计入明细、小计、合计）
					if not (stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽'):
						continue
					# 特殊处理：张佳丽固定按药师的奖励规则计算
					if stats['profession'] == '药师' or stats['reporter_name'] == '张佳丽':
						# 药师按照阶梯式计算
						reward_一般 = calculate_tiered_reward(stats['一般'], 50)
						reward_严重 = calculate_tiered_reward(stats['严重'], 70)
						reward_新的一般 = calculate_tiered_reward(stats['新的_一般'], 100)
						reward_新的严重 = calculate_tiered_reward(stats['新的_严重'], 150)

						individual_total = reward_一般 + reward_严重 + reward_新的一般 + reward_新的严重
					else:
						# 非药师按300元/例
						total_cases = stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
						individual_total = total_cases * 300
						reward_一般 = stats['一般'] * 300 if stats['一般'] > 0 else 0
						reward_严重 = stats['严重'] * 300 if stats['严重'] > 0 else 0
						reward_新的一般 = stats['新的_一般'] * 300 if stats['新的_一般'] > 0 else 0
						reward_新的严重 = stats['新的_严重'] * 300 if stats['新的_严重'] > 0 else 0

					reward_excel_data.append({
						'科室': stats['department'],
						'职业': stats['profession'],
						'姓名': stats['reporter_name'],
						'总数量': stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重'],
						'一般': stats['一般'],
						'奖励/元': int(reward_一般),
						'严重': stats['严重'],
						'奖励/元.1': int(reward_严重),
						'新的一般': stats['新的_一般'],
						'奖励/元.2': int(reward_新的一般),
						'新的严重': stats['新的_严重'],
						'奖励/元.3': int(reward_新的严重),
						'个人奖励合计/元': int(individual_total)
					})
					
					# 计算科室小计
					dept = stats['department']
					if dept not in reward_department_subtotals:
						reward_department_subtotals[dept] = {
							'总数量': 0,
							'一般': 0,
							'奖励_一般': 0,
							'严重': 0,
							'奖励_严重': 0,
							'新的_一般': 0,
							'奖励_新的一般': 0,
							'新的_严重': 0,
							'奖励_新的严重': 0,
							'个人奖励合计': 0
						}

					reward_department_subtotals[dept]['总数量'] += stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
					reward_department_subtotals[dept]['一般'] += stats['一般']
					reward_department_subtotals[dept]['奖励_一般'] += int(reward_一般)
					reward_department_subtotals[dept]['严重'] += stats['严重']
					reward_department_subtotals[dept]['奖励_严重'] += int(reward_严重)
					reward_department_subtotals[dept]['新的_一般'] += stats['新的_一般']
					reward_department_subtotals[dept]['奖励_新的一般'] += int(reward_新的一般)
					reward_department_subtotals[dept]['新的_严重'] += stats['新的_严重']
					reward_department_subtotals[dept]['奖励_新的严重'] += int(reward_新的严重)
					reward_department_subtotals[dept]['个人奖励合计'] += int(individual_total)
					
					# 累计总数
					reward_total_stats['总数量'] += stats['一般'] + stats['严重'] + stats['新的_一般'] + stats['新的_严重']
					reward_total_stats['一般'] += stats['一般']
					reward_total_stats['奖励_一般'] += int(reward_一般)
					reward_total_stats['严重'] += stats['严重']
					reward_total_stats['奖励_严重'] += int(reward_严重)
					reward_total_stats['新的_一般'] += stats['新的_一般']
					reward_total_stats['奖励_新的一般'] += int(reward_新的一般)
					reward_total_stats['新的_严重'] += stats['新的_严重']
					reward_total_stats['奖励_新的严重'] += int(reward_新的严重)
					reward_total_stats['个人奖励合计'] += int(individual_total)
				
				# 按科室、职业、姓名排序
				reward_excel_data.sort(key=lambda x: (x['科室'], x['职业'], x['姓名']))
				
				# 创建完整的表格数据（包含小计和合计）
				reward_final_data = []
				current_dept = None
				
				for row in reward_excel_data:
					# 如果是新科室，先添加上一个科室的小计
					if current_dept and current_dept != row['科室']:
						if current_dept in reward_department_subtotals:
							subtotal = reward_department_subtotals[current_dept]
							reward_final_data.append({
								'科室': '小计',
								'职业': '',
								'姓名': '',
								'总数量': subtotal['一般'] + subtotal['严重'] + subtotal['新的_一般'] + subtotal['新的_严重'],
								'一般': subtotal['一般'],
								'奖励/元': subtotal['奖励_一般'],
								'严重': subtotal['严重'],
								'奖励/元.1': subtotal['奖励_严重'],
								'新的一般': subtotal['新的_一般'],
								'奖励/元.2': subtotal['奖励_新的一般'],
								'新的严重': subtotal['新的_严重'],
								'奖励/元.3': subtotal['奖励_新的严重'],
								'个人奖励合计/元': subtotal['个人奖励合计']
							})
					
					reward_final_data.append(row)
					current_dept = row['科室']
				
				# 添加最后一个科室的小计
				if current_dept and current_dept in reward_department_subtotals:
					subtotal = reward_department_subtotals[current_dept]
					reward_final_data.append({
						'科室': '小计',
						'职业': '',
						'姓名': '',
						'总数量': subtotal['一般'] + subtotal['严重'] + subtotal['新的_一般'] + subtotal['新的_严重'],
						'一般': subtotal['一般'],
						'奖励/元': subtotal['奖励_一般'],
						'严重': subtotal['严重'],
						'奖励/元.1': subtotal['奖励_严重'],
						'新的一般': subtotal['新的_一般'],
						'奖励/元.2': subtotal['奖励_新的一般'],
						'新的严重': subtotal['新的_严重'],
						'奖励/元.3': subtotal['奖励_新的严重'],
						'个人奖励合计/元': subtotal['个人奖励合计']
					})
				
				# 添加合计行
				reward_final_data.append({
					'科室': '合计',
					'职业': '',
					'姓名': '',
					'总数量': '',
					'一般': '',
					'奖励/元': '',
					'严重': '',
					'奖励/元.1': '',
					'新的一般': '',
					'奖励/元.2': '',
					'新的严重': '',
					'奖励/元.3': '',
					'个人奖励合计/元': reward_total_stats['个人奖励合计']
				})
				
				df_reward = pd.DataFrame(reward_final_data)
				df_reward.to_excel(writer, sheet_name='奖励计算-药学', index=False, header=False, startrow=2)
				worksheet_reward = writer.sheets['奖励计算-药学']
				apply_reward_template(worksheet_reward)
				
				# 添加奖励计算图表 - 使用matplotlib
				try:
					import matplotlib
					matplotlib.use('Agg')
					import matplotlib.pyplot as plt
					
					plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
					plt.rcParams['axes.unicode_minus'] = False
					
					from openpyxl.drawing.image import Image
					
					# 准备个人奖励数据
					person_rewards = [(row['姓名'], row['个人奖励合计/元']) 
									  for row in reward_excel_data if row['姓名'] and row['科室'] not in ['小计', '合计']]
					person_rewards.sort(key=lambda x: x[1], reverse=True)
					top_persons = person_rewards[:10]
					
					# 1. 个人奖励TOP10柱状图
					if top_persons:
						fig1, ax1 = plt.subplots(figsize=(10, 6))
						names = [p[0] for p in top_persons]
						rewards = [p[1] for p in top_persons]
						bars = ax1.barh(names[::-1], rewards[::-1], color='#E6A23C')
						ax1.set_title('个人奖励TOP10', fontsize=14, fontweight='bold')
						ax1.set_xlabel('奖励金额(元)')
						for bar, val in zip(bars, rewards[::-1]):
							ax1.text(bar.get_width() + 10, bar.get_y() + bar.get_height()/2, f'¥{val}', va='center', fontsize=10)
						plt.tight_layout()
						
						chart1_image = BytesIO()
						plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
						chart1_image.seek(0)
						plt.close()
						
						img1 = Image(chart1_image)
						img1.width = 500
						img1.height = 300
						worksheet_reward.add_image(img1, 'O2')
					
					# 2. 奖励类型占比饼图
					type_rewards = [
						('一般', reward_total_stats['奖励_一般']),
						('严重', reward_total_stats['奖励_严重']),
						('新的+一般', reward_total_stats['奖励_新的一般']),
						('新的+严重', reward_total_stats['奖励_新的严重'])
					]
					type_rewards = [(t, v) for t, v in type_rewards if v > 0]
					
					if type_rewards:
						fig2, ax2 = plt.subplots(figsize=(10, 7))
						labels = [f"{t[0]}: ¥{t[1]}" for t in type_rewards]
						values = [t[1] for t in type_rewards]
						colors = ['#67C23A', '#F56C6C', '#409EFF', '#E6A23C']
						draw_pie_with_labels(ax2, values, labels, colors, '奖励类型占比')
						
						chart2_image = BytesIO()
						plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
						chart2_image.seek(0)
						plt.close()
						
						img2 = Image(chart2_image)
						img2.width = 450
						img2.height = 350
						worksheet_reward.add_image(img2, 'O20')
						
				except Exception as chart_err:
					print(f"生成奖励计算图表时出错: {chart_err}")
				
			except Exception as e:
				print(f"奖励计算导出错误: {str(e)}")
			
			# 6. 发生不良反应的药品汇总数据 - 复用export_drug_summary的核心逻辑
			try:
				# 按药品统计
				drug_stats = {}
				
				for record in all_records:
					drug_name = record.generic_name or "未知药品"
					manufacturer = record.manufacturer or "未知厂家"
					
					key = f"{drug_name}|{manufacturer}"
					
					if key not in drug_stats:
						drug_stats[key] = {
							'drug_name': drug_name,
							'manufacturer': manufacturer,
							'一般': 0,
							'严重': 0,
							'total': 0
						}
					
					if record.severity == '一般':
						drug_stats[key]['一般'] += 1
					elif record.severity == '严重':
						drug_stats[key]['严重'] += 1
					
					drug_stats[key]['total'] += 1
				
				# 转换为列表并按合计数量排序
				drug_list = list(drug_stats.values())
				drug_list.sort(key=lambda x: x['total'], reverse=True)
				
				# 创建Excel数据
				drug_excel_data = []
				for i, drug in enumerate(drug_list, 1):
					drug_excel_data.append({
						'序号': i,
						'通用名称': drug['drug_name'],
						'生产厂家': drug['manufacturer'],
						'严重': drug['严重'],
						'一般': drug['一般'],
						'合计': drug['total']
					})
				
				# 添加合计行
				if drug_excel_data:
					total_severe = sum(item['严重'] for item in drug_excel_data)
					total_general = sum(item['一般'] for item in drug_excel_data)
					drug_excel_data.append({
						'序号': '',
						'通用名称': '合计',
						'生产厂家': '',
						'严重': total_severe,
						'一般': total_general,
						'合计': total_severe + total_general
					})

				df_drug_summary = pd.DataFrame(drug_excel_data)
				df_drug_summary.to_excel(writer, sheet_name='发生不良反应的药品汇总', index=False)
				df_reaction_summary = pd.DataFrame(build_reaction_summary_excel_data(all_records))
				df_reaction_summary.to_excel(writer, sheet_name='不良反应汇总分析', index=False)
				add_reaction_summary_charts(writer, all_records)

				# 添加药品汇总图表 - 使用matplotlib
				try:
					import matplotlib
					matplotlib.use('Agg')
					import matplotlib.pyplot as plt
					import numpy as np
					
					plt.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'Microsoft YaHei', 'DejaVu Sans']
					plt.rcParams['axes.unicode_minus'] = False
					
					from openpyxl.drawing.image import Image
					worksheet_drug = writer.sheets['发生不良反应的药品汇总']
					
					# 1. 前十名药品柱状图
					top_10_drugs = drug_list[:10]
					if top_10_drugs:
						fig1, ax1 = plt.subplots(figsize=(12, 7))
						drug_names = [d['drug_name'][:15] + '..' if len(d['drug_name']) > 15 else d['drug_name'] for d in top_10_drugs]
						general_counts = [d['一般'] for d in top_10_drugs]
						severe_counts = [d['严重'] for d in top_10_drugs]
						
						x = np.arange(len(drug_names))
						width = 0.35
						bars1 = ax1.barh(x - width/2, general_counts[::-1], width, label='一般', color='#67C23A')
						bars2 = ax1.barh(x + width/2, severe_counts[::-1], width, label='严重', color='#F56C6C')
						
						ax1.set_yticks(x)
						ax1.set_yticklabels(drug_names[::-1], fontsize=9)
						ax1.set_title('药品不良反应前十名', fontsize=14, fontweight='bold')
						ax1.set_xlabel('数量')
						ax1.legend()
						plt.tight_layout()
						
						chart1_image = BytesIO()
						plt.savefig(chart1_image, format='png', dpi=150, bbox_inches='tight')
						chart1_image.seek(0)
						plt.close()
						
						img1 = Image(chart1_image)
						img1.width = 600
						img1.height = 350
						worksheet_drug.add_image(img1, 'H2')
					
					# 2. 药品分布饼图
					top_10_for_pie = drug_list[:10]
					if top_10_for_pie:
						fig2, ax2 = plt.subplots(figsize=(12, 9))
						labels = [f"{d['drug_name']}: {d['total']}例" for d in top_10_for_pie]
						values = [d['total'] for d in top_10_for_pie]
						other_total = sum(d['total'] for d in drug_list[10:])
						if other_total > 0:
							labels.append(f'其他: {other_total}例')
							values.append(other_total)
						
						colors = ['#409EFF', '#67C23A', '#E6A23C', '#F56C6C', '#909399', '#606266', '#E91E63', '#9C27B0', '#3F51B5', '#00BCD4', '#CCCCCC']
						draw_pie_with_labels(ax2, values, labels, colors, '药品不良反应分布')
						
						chart2_image = BytesIO()
						plt.savefig(chart2_image, format='png', dpi=150, bbox_inches='tight')
						chart2_image.seek(0)
						plt.close()
						
						img2 = Image(chart2_image)
						img2.width = 600
						img2.height = 450
						worksheet_drug.add_image(img2, 'H24')
						
				except Exception as chart_err:
					print(f"生成药品汇总图表时出错: {chart_err}")
				
			except Exception as e:
				print(f"药品汇总导出错误: {str(e)}")
			
		
		output.seek(0)
		
		# 生成文件名
		filename = f"药品不良反应数据分析汇总_{get_month_range_str(start_month, end_month)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		return send_file(
			output,
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
			as_attachment=True,
			download_name=filename
		)
		
	except Exception as e:
		return jsonify({"message": f"合并导出数据失败: {str(e)}"}), 500

@data_bp.route("/reports/batch-delete", methods=["POST"])
@require_auth
def batch_delete_reports():
	"""批量删除报告"""
	try:
		data = request.get_json()
		report_ids = data.get('ids', [])

		if not report_ids:
			return jsonify({"message": "请选择要删除的报告"}), 400

		# 删除记录
		deleted_count = AdverseReactionReport.query.filter(
			AdverseReactionReport.id.in_(report_ids)
		).delete(synchronize_session=False)

		db.session.commit()

		return jsonify({
			"message": f"成功删除 {deleted_count} 条记录",
			"count": deleted_count
		})

	except Exception as e:
		db.session.rollback()
		return jsonify({"message": f"批量删除失败: {str(e)}"}), 500

@data_bp.route("/reports/clear-all", methods=["POST"])
@require_auth
def clear_all_reports():
	"""清空所有报告数据"""
	try:
		# 统计总数
		total_count = AdverseReactionReport.query.count()

		# 删除所有记录
		AdverseReactionReport.query.delete()

		# 清空导入历史
		ImportHistory.query.delete()

		db.session.commit()

		return jsonify({
			"message": f"成功清空数据库，共删除 {total_count} 条记录",
			"count": total_count
		})

	except Exception as e:
		db.session.rollback()
		return jsonify({"message": f"清空数据库失败: {str(e)}"}), 500

@data_bp.route("/analysis/drug-summary-charts", methods=["GET"])
@require_auth
def get_drug_summary_charts():
	"""获取药品汇总图表数据"""
	try:
		# 获取筛选参数
		start_month = request.args.get('startMonth', '')
		end_month = request.args.get('endMonth', '')

		query = AdverseReactionReport.query
		query = apply_month_filter(query, start_month, end_month)
		
		# 先过滤掉"并用"的数据，只保留"怀疑"的数据
		filtered_query = query.filter(AdverseReactionReport.suspect_concurrent == '怀疑')
		
		# 然后按报告编码去重（取每个报告编码的第一条记录）
		subquery = filtered_query.with_entities(
			AdverseReactionReport.report_code,
			db.func.min(AdverseReactionReport.id).label('min_id')
		).group_by(AdverseReactionReport.report_code).subquery()
		
		deduplicated_query = db.session.query(AdverseReactionReport).join(
			subquery, AdverseReactionReport.id == subquery.c.min_id
		)
		
		# 获取所有去重后的记录
		all_records = deduplicated_query.all()
		
		# 按药品统计
		drug_stats = {}
		
		for record in all_records:
			# 获取药品信息
			drug_name = record.generic_name or "未知药品"
			manufacturer = record.manufacturer or "未知厂家"
			
			# 创建唯一键
			key = f"{drug_name}|{manufacturer}"
			
			if key not in drug_stats:
				drug_stats[key] = {
					'drug_name': drug_name,
					'manufacturer': manufacturer,
					'一般': 0,
					'严重': 0,
					'total': 0
				}
			
			# 统计严重程度
			if record.severity == '一般':
				drug_stats[key]['一般'] += 1
			elif record.severity == '严重':
				drug_stats[key]['严重'] += 1
			
			drug_stats[key]['total'] += 1
		
		# 转换为列表并按合计数量排序
		drug_list = list(drug_stats.values())
		drug_list.sort(key=lambda x: x['total'], reverse=True)
		
		# 准备前十名柱状图数据（不省略名称）
		top_10_drugs = drug_list[:10]
		bar_chart_data = []

		for drug in top_10_drugs:
			bar_chart_data.append({
				'drug_name': drug['drug_name'],
				'manufacturer': drug['manufacturer'],
				'一般': drug['一般'],
				'严重': drug['严重'],
				'total': drug['total']
			})

		# 准备饼状图数据（显示前10名，其余合并为"其他"，不省略名称）
		pie_chart_data = []
		top_10_for_pie = drug_list[:10]
		other_total = sum(drug['total'] for drug in drug_list[10:])

		for drug in top_10_for_pie:
			pie_chart_data.append({
				'name': drug['drug_name'],
				'value': drug['total']
			})

		if other_total > 0:
			pie_chart_data.append({
				'name': '其他',
				'value': other_total
			})
		
		return jsonify({
			'bar_chart': bar_chart_data,
			'pie_chart': pie_chart_data
		})
		
	except Exception as e:
		return jsonify({"message": f"获取药品汇总图表数据失败: {str(e)}"}), 500
