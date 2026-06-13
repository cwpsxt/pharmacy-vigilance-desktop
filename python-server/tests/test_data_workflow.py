import os
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

os.environ["SKIP_LICENSE"] = "1"

import app as app_module
from app.db import db
from app.models import AdverseReactionReport


class DataWorkflowTest(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_user_data_dir = app_module.get_user_data_dir
        app_module.get_user_data_dir = lambda: self.tmp.name
        self.app = app_module.create_app()
        self.client = self.app.test_client()
        response = self.client.post("/api/auth/login", json={"username": "admin", "password": "admin"})
        self.assertEqual(response.status_code, 200)
        self.token = response.get_json()["token"]
        self.headers = {"Authorization": f"Bearer {self.token}"}

    def tearDown(self):
        app_module.get_user_data_dir = self.old_user_data_dir
        self.tmp.cleanup()

    def _xlsx_bytes(self, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "报告信息"
        for row in rows:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _upload(self, rows, filename="报告导出(14)(1).xlsx"):
        data = {"file": (self._xlsx_bytes(rows), filename)}
        return self.client.post(
            "/api/data/upload",
            data=data,
            headers=self.headers,
            content_type="multipart/form-data",
        )

    def test_import_accepts_center_export_with_preamble_row(self):
        headers = [
            "报告表编码",
            "报告类型-新的",
            "报告类型-严重程度",
            "病历号/门诊号",
            "怀疑/并用",
            "通用名称",
            "生产厂家",
            "不良反应名称",
            "报告人职业",
            "报告人签名",
            "国家中心接收时间",
        ]
        response = self._upload([
            ["国家中心导出文件", "", "", "", "", "", "", "", "", "", ""],
            headers,
            ["ADR-001", "新的", "严重", "M001", "怀疑", "阿莫西林", "示例药厂", "皮疹", "药师", "张三", "2024-08-01 08:00:00"],
        ])
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertEqual(payload["success_records"], 1)
        self.assertEqual(payload["failed_records"], 0)

    def test_import_history_delete_removes_batch_reports(self):
        headers = ["报告表编码", "报告类型-新的", "报告类型-严重程度", "病历号/门诊号", "怀疑/并用", "通用名称", "生产厂家", "不良反应名称", "报告人职业", "报告人签名", "国家中心接收时间"]
        upload = self._upload([headers, ["ADR-DEL", "", "一般", "M002", "怀疑", "头孢", "示例药厂", "恶心", "药师", "李四", "2024-08-02 08:00:00"]])
        self.assertEqual(upload.status_code, 200)
        history = self.client.get("/api/data/import-history", headers=self.headers).get_json()
        response = self.client.delete(f"/api/data/import-history/{history[0]['id']}", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        with self.app.app_context():
            self.assertEqual(AdverseReactionReport.query.count(), 0)

    def test_drug_summary_export_includes_reaction_summary_sheet(self):
        headers = ["报告表编码", "报告类型-新的", "报告类型-严重程度", "病历号/门诊号", "怀疑/并用", "通用名称", "生产厂家", "不良反应名称", "报告人职业", "报告人签名", "国家中心接收时间"]
        upload = self._upload([headers, ["ADR-SUM", "", "一般", "M003", "怀疑", "头孢", "示例药厂", "恶心；呕吐", "药师", "王五", "2024-08-03 08:00:00"]])
        self.assertEqual(upload.status_code, 200)
        response = self.client.get("/api/data/analysis/export-drug-summary", headers=self.headers)
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.data), read_only=True)
        self.assertIn("不良反应汇总分析", wb.sheetnames)
        ws = wb["不良反应汇总分析"]
        self.assertEqual([cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))[:5]], ["序号", "不良反应名称", "严重", "一般", "合计"])

    def test_report_details_and_reward_exports_use_template_headers(self):
        headers = ["报告表编码", "报告类型-新的", "报告类型-严重程度", "病历号/门诊号", "怀疑/并用", "通用名称", "生产厂家", "不良反应名称", "报告人职业", "报告人签名", "国家中心接收时间"]
        upload = self._upload([headers, ["ADR-TPL", "新的", "严重", "M004", "怀疑", "头孢", "示例药厂", "皮疹", "药师", "赵六", "2024-08-04 08:00:00"]])
        self.assertEqual(upload.status_code, 200)

        details_response = self.client.get("/api/data/analysis/export-report-details", headers=self.headers)
        self.assertEqual(details_response.status_code, 200)
        details_wb = load_workbook(BytesIO(details_response.data), read_only=False)
        details_ws = details_wb["上报明细"]
        self.assertEqual([details_ws.cell(1, c).value for c in range(1, 9)], ["科室", "职业", "姓名", "总计", "报告类型", None, None, None])
        self.assertEqual(sorted(str(r) for r in details_ws.merged_cells.ranges), ["A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:H1"])

        reward_response = self.client.get("/api/data/analysis/export-reward-calculation", headers=self.headers)
        self.assertEqual(reward_response.status_code, 200)
        reward_wb = load_workbook(BytesIO(reward_response.data), read_only=False)
        reward_ws = reward_wb["奖励计算"]
        self.assertEqual([reward_ws.cell(1, c).value for c in range(1, 13)], ["科室", "职业", "姓名", "一般", None, "严重", None, "新的一般", None, "新的严重", None, "个人奖励合计/元"])
        self.assertEqual(sorted(str(r) for r in reward_ws.merged_cells.ranges), ["A1:A2", "B1:B2", "C1:C2", "D1:E1", "F1:G1", "H1:I1", "J1:K1", "L1:L2"])

    def test_desktop_pages_expose_delete_actions_and_correct_import_count(self):
        reports_html = (ROOT.parent / "electron/pages/views/reports.html").read_text(encoding="utf-8")
        import_html = (ROOT.parent / "electron/pages/views/import.html").read_text(encoding="utf-8")
        self.assertIn("/api/data/reports/batch-delete", reports_html)
        self.assertIn("/api/data/reports/clear-all", reports_html)
        self.assertIn("/api/data/import-history/${item.id}", import_html)
        self.assertIn("data.success_records", import_html)
        self.assertNotIn("data.imported_count", import_html)


if __name__ == "__main__":
    unittest.main()
